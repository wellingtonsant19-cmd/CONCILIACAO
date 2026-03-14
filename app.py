import streamlit as st
import pandas as pd
import numpy as np
from itertools import combinations
import time
import re as _re
import unicodedata as _unicodedata
import io

# ============================================================
# CONFIGURAÇÃO DA PÁGINA
# ============================================================

st.set_page_config(
    page_title="Motor de Conciliação",
    page_icon="⚖️",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;700&family=IBM+Plex+Sans:wght@400;600;800&display=swap');

html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
code, .mono { font-family: 'IBM Plex Mono', monospace; }

.header-box {
    background: linear-gradient(135deg, #0f2044 0%, #0a1628 100%);
    border: 1px solid #1e3a5f;
    border-radius: 12px;
    padding: 24px 32px;
    margin-bottom: 28px;
}
.stat-box {
    background: #0a1628;
    border: 1px solid #1e293b;
    border-radius: 10px;
    padding: 16px 20px;
    text-align: center;
}
.stat-label { font-size: 10px; color: #64748b; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 4px; }
.stat-value { font-size: 22px; font-weight: 800; color: #e2e8f0; font-family: 'IBM Plex Mono', monospace; }
.stat-value.green  { color: #4ade80; }
.stat-value.yellow { color: #f59e0b; }
.stat-value.red    { color: #f87171; }

.alert-retencao {
    background: #ef444410; border: 1px solid #ef444430;
    border-radius: 8px; padding: 14px 18px;
    color: #fca5a5; font-size: 13px; margin-top: 12px;
}
.alert-multi {
    background: #f59e0b10; border: 1px solid #f59e0b30;
    border-radius: 8px; padding: 12px 16px;
    color: #fcd34d; font-size: 13px; margin-bottom: 16px;
}
.alert-escopo {
    background: #8b5cf610; border: 1px solid #8b5cf630;
    border-radius: 8px; padding: 12px 16px;
    color: #c4b5fd; font-size: 13px; margin-bottom: 16px;
}
.alert-success {
    background: #22c55e10; border: 1px solid #22c55e30;
    border-radius: 8px; padding: 12px 16px;
    color: #86efac; font-size: 13px; margin-bottom: 16px;
}
.badge {
    display: inline-block; border-radius: 4px;
    padding: 2px 8px; font-size: 11px; font-weight: 700; letter-spacing: 0.04em;
}
.badge-green  { background: #22c55e20; color: #4ade80;  border: 1px solid #22c55e40; }
.badge-yellow { background: #f59e0b20; color: #fbbf24;  border: 1px solid #f59e0b40; }
.badge-red    { background: #ef444420; color: #f87171;  border: 1px solid #ef444440; }
.badge-blue   { background: #3b82f620; color: #60a5fa;  border: 1px solid #3b82f640; }
.badge-purple { background: #8b5cf620; color: #c4b5fd;  border: 1px solid #8b5cf640; }
.badge-gray   { background: #33415520; color: #94a3b8;  border: 1px solid #33415540; }

div[data-testid="stDataFrame"] { border: 1px solid #1e293b; border-radius: 8px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# CONFIGURAÇÕES DO ALGORITMO
# ============================================================

TOLERANCIA     = 0.01
MAX_NOTAS      = 30
TOP_RESULTADOS = 10

# ============================================================
# LEITURA E NORMALIZAÇÃO DA PLANILHA
# ============================================================

COLUNAS = {
    "nome":         ["NOME CLIENTE", "NOME"],
    "cnpj":         ["CNPJ"],
    "cidade":       ["CIDADE"],
    "uf":           ["UF"],
    "nf":           ["NF", "NR NFEM"],
    "num_doc":      ["DOC", "NUM DOC MATERA"],
    "rbase":        ["RBASE RAIZ", "RBASE"],
    "valor_titulo": ["BRUTO", "VLR TITULO", "VALOR TITULO"],
    "saldo":        ["SALDO", "VLR SALDO", "VALOR SALDO"],
    "ir":           ["IR", "VLR RETIDO"],
    "iss":          ["ISS", "ISS RETIDO"],
    "venc":         ["VENCIMENTO", "VENC", "DT VENCIMENTO"],
    "atraso":       ["ATRASO"],
}

def _get_col(df, candidates):
    cols_upper = {c.strip().upper(): c for c in df.columns}
    for cand in candidates:
        if cand.upper() in cols_upper:
            return cols_upper[cand.upper()]
    return None

@st.cache_data(show_spinner=False)
def carregar_planilha(file_bytes, file_name):
    df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = df.columns.str.strip()
    renomear = {}
    for destino, candidates in COLUNAS.items():
        col = _get_col(df, candidates)
        if col:
            renomear[col] = destino
    df.rename(columns=renomear, inplace=True)
    for col in ["valor_titulo", "saldo", "ir", "iss", "atraso"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(2)
    if "cnpj" in df.columns:
        df["cnpj_limpo"] = df["cnpj"].astype(str).str.replace(r"\D", "", regex=True)
    df.index = range(len(df))
    return df

# ============================================================
# ALGORITMO — CAMADA 1: PRÉ-PROCESSAMENTO
# ============================================================

def preprocessar(valores, alvo):
    limite = round(alvo + TOLERANCIA, 2)
    filtrado = {k: v for k, v in valores.items() if 0 < round(v, 2) <= limite}
    return dict(sorted(filtrado.items(), key=lambda x: x[1], reverse=True))

# ============================================================
# ALGORITMO — CAMADA 2: GREEDY DESCENDING
# ============================================================

def greedy(valores, alvo, max_notas):
    alvo_r = round(alvo, 2)
    soma, combo = 0.0, []
    for chave, val in valores.items():
        if len(combo) >= max_notas: break
        if round(soma + val, 2) <= round(alvo_r + TOLERANCIA, 2):
            soma = round(soma + val, 2)
            combo.append(chave)
        if abs(soma - alvo_r) <= TOLERANCIA:
            return (round(abs(soma - alvo_r), 2), tuple(combo))
    return None

# ============================================================
# ALGORITMO — CAMADA 3: MEET-IN-THE-MIDDLE (adaptativo)
# ============================================================

def _somas_grupo(items, max_n, cap=None):
    dp = {0: [()]}
    MAX_DP_SIZE = 50_000  # guard: abort if DP table grows too large
    for chave, val in items:
        new_e = {}
        for sa, combos in dp.items():
            nv = sa + val
            novos = [c + (chave,) for c in combos if len(c) < max_n]
            if novos:
                if nv not in new_e: new_e[nv] = []
                new_e[nv].extend(novos)
        for nv, novos in new_e.items():
            if nv not in dp:
                dp[nv] = novos[:cap] if cap else novos
            else:
                merged = dp[nv] + novos
                if cap and len(merged) > cap:
                    merged.sort(key=len)
                    merged = merged[:cap]
                dp[nv] = merged
        if len(dp) > MAX_DP_SIZE:
            break
    return dp

def meet_in_the_middle(valores, alvo, max_notas):
    n = len(valores)
    # Cap adaptativo — conservador para evitar memory explosion
    if   n <= 20:  cap = None   # exato — seguro
    elif n <= 35:  cap = 5
    elif n <= 60:  cap = 2
    else:          cap = 1

    alvo_cents = round(round(alvo, 2) * 100)
    tol_cents  = round(TOLERANCIA * 100)
    items = [(k, round(round(v, 2) * 100)) for k, v in valores.items() if round(v, 2) > 0]
    if not items: return []

    mid   = len(items) // 2
    left, right = items[:mid], items[mid:]
    max_l = max(1, max_notas * len(left)  // len(items))
    max_r = max(1, max_notas * len(right) // len(items))

    somas_l = _somas_grupo(left,  max_l, cap)
    somas_r = _somas_grupo(right, max_r, cap)

    resultados, vistas = [], set()
    for soma_l, combos_l in somas_l.items():
        needed = alvo_cents - soma_l
        for delta in range(-tol_cents, tol_cents + 1):
            soma_r = needed + delta
            if soma_r not in somas_r: continue
            for cl in combos_l:
                for cr in somas_r[soma_r]:
                    merged = frozenset(cl + cr)
                    if len(merged) > max_notas or merged in vistas: continue
                    vistas.add(merged)
                    total = soma_l + soma_r
                    dif   = round(abs(alvo_cents - total) / 100, 2)
                    resultados.append((dif, tuple(sorted(merged))))

    resultados.sort(key=lambda x: (x[0], len(x[1])))
    return resultados[:TOP_RESULTADOS * 3]

def fix1_busca(valores, alvo, max_notas):
    resultados = []
    thr = alvo / max(max_notas, 1)
    sd  = sorted(valores.items(), key=lambda x: x[1], reverse=True)
    ancoras = sorted([(k, v) for k, v in sd if thr < v < alvo - TOLERANCIA],
                     key=lambda x: abs(x[1] - alvo / 2))
    vistas = set()
    for k1, v1 in ancoras:
        resto = round(alvo - v1, 2)
        sub = dict(sorted(
            {k: v for k, v in valores.items() if k != k1 and 0 < v <= resto + TOLERANCIA}.items(),
            key=lambda x: x[1], reverse=True
        )[:50])
        if not sub: continue
        for _, combo in meet_in_the_middle(sub, resto, max_notas - 1)[:5]:
            full = tuple(sorted((k1,) + combo))
            key  = frozenset(full)
            if key not in vistas:
                vistas.add(key)
                soma = round(v1 + sum(valores.get(k, 0) for k in combo), 2)
                resultados.append((round(abs(soma - alvo), 2), full))
        if resultados and resultados[0][0] <= TOLERANCIA:
            break
    return resultados

# ============================================================
# ALGORITMO — CAMADA 4: SOLVER MILP (fallback)
# ============================================================

def solver_milp(valores, alvo, max_notas):
    try:
        from scipy.optimize import milp, LinearConstraint, Bounds
        keys = list(valores.keys())
        vals = np.array([round(valores[k], 2) for k in keys])
        n    = len(keys)
        c    = np.ones(n)
        tol  = TOLERANCIA

        # Restrição 1: soma = alvo (±tolerância)
        constraints = [LinearConstraint(vals.reshape(1, -1), lb=alvo - tol, ub=alvo + tol)]

        # Restrição 2: máximo de notas
        constraints.append(LinearConstraint(np.ones((1, n)), lb=0, ub=max_notas))

        # Restrição 3: no máximo 1 variante por nota (idx antes do "|")
        note_groups = {}
        for i, k in enumerate(keys):
            note_idx = k.split("|")[0]
            note_groups.setdefault(note_idx, []).append(i)

        for note_idx, indices in note_groups.items():
            if len(indices) > 1:
                row = np.zeros(n)
                for i in indices:
                    row[i] = 1
                constraints.append(LinearConstraint(row.reshape(1, -1), lb=0, ub=1))

        bounds = Bounds(lb=np.zeros(n), ub=np.ones(n))
        res = milp(c, constraints=constraints, integrality=np.ones(n), bounds=bounds)
        if res.success and res.x is not None:
            xi    = np.round(res.x).astype(int)
            combo = tuple(keys[i] for i in range(n) if xi[i] == 1)
            soma  = round(sum(valores[k] for k in combo), 2)
            dif   = round(abs(soma - round(alvo, 2)), 2)
            return [(dif, combo)]
    except Exception:
        pass
    return []

# ============================================================
# ALGORITMO — CAMADA 5: RANKING INTELIGENTE
# ============================================================

def _score(combo, valores, df_cli, alvo):
    alvo_r  = round(alvo, 2)
    soma    = round(sum(round(valores.get(k, 0), 2) for k in combo), 2)
    dif     = round(abs(soma - alvo_r), 2)
    n_notas = len(combo)
    rbases, atrasos = set(), []
    for chave in combo:
        idx_str, _ = chave.split("|")
        try:
            row = df_cli.loc[int(idx_str)]
            if "rbase" in df_cli.columns:
                rbases.add(str(row.get("rbase", "")))
            try:
                atrasos.append(float(row.get("atraso", 0) or 0))
            except Exception:
                atrasos.append(0)
        except Exception:
            pass
    media_atraso = float(np.mean(atrasos)) if atrasos else 0
    return (dif, n_notas, len(rbases), -media_atraso)

def _combo_sem_nf_duplicada(combo):
    indices = [chave.split("|")[0] for chave in combo]
    return len(indices) == len(set(indices))

def ranquear(brutos, valores, df_cli, alvo):
    vistas_idx, com_score = set(), []
    for dif, combo in brutos:
        if not _combo_sem_nf_duplicada(combo): continue
        idx_set = frozenset(chave.split("|")[0] for chave in combo)
        if idx_set in vistas_idx: continue
        vistas_idx.add(idx_set)
        score = _score(combo, valores, df_cli, alvo)
        com_score.append((score, combo))
    com_score.sort(key=lambda x: x[0])
    return [(round(s[0], 2), c) for s, c in com_score[:TOP_RESULTADOS]]

# ============================================================
# ORQUESTRADOR
# ============================================================

def gerar_opcoes_nota(row):
    s   = round(float(row.get("saldo", 0)), 2)
    ir  = round(float(row.get("ir",    0)), 2)
    iss = round(float(row.get("iss",   0)), 2)
    ops = [("saldo", s)]
    if ir  != 0: ops.append(("saldo_ir",      round(s + ir,       2)))
    if iss != 0: ops.append(("saldo_iss",     round(s + iss,      2)))
    if ir != 0 and iss != 0:
        ops.append(("saldo_ir_iss", round(s + ir + iss, 2)))
    elif ir != 0 or iss != 0:
        liq = round(s + ir + iss, 2)
        # Só adiciona se for diferente do que já existe
        vals_existentes = {v for _, v in ops}
        if liq not in vals_existentes:
            ops.append(("saldo_ir_iss", liq))
    return ops

def buscar_combinacoes(valores_raw, alvo, df_cli, max_notas=MAX_NOTAS):
    valores = preprocessar(valores_raw, alvo)
    if not valores: return [], {}

    n_filtradas = len(valores)
    brutos = []

    g = greedy(valores, alvo, max_notas)
    if g: brutos.append(g)

    mitm = meet_in_the_middle(valores, alvo, max_notas)
    brutos.extend(mitm)

    if n_filtradas > 50 and not (brutos and brutos[0][0] <= TOLERANCIA):
        fix1 = fix1_busca(valores, alvo, max_notas)
        brutos.extend(fix1)

    # MILP roda sempre que não tiver match exato (não só quando brutos vazio)
    tem_exato = any(dif <= TOLERANCIA for dif, _ in brutos)
    if not tem_exato:
        milp_res = solver_milp(valores, alvo, max_notas)
        brutos.extend(milp_res)

    if not brutos: return [], {"pre": n_filtradas, "camada": "Nenhuma"}

    camada = ("Greedy"              if g and g[0] <= TOLERANCIA
              else "MILP"           if not g and not mitm
              else "MITM + Fix-1"   if mitm and n_filtradas > 50
              else "Meet-in-the-Middle" if mitm
              else "MILP")
    resultado = ranquear(brutos, valores, df_cli, alvo)
    return resultado, {"pre": n_filtradas, "camada": camada}

def _montar_df_valores(df_sub):
    valores = {}
    for idx, row in df_sub.iterrows():
        for nome, valor in gerar_opcoes_nota(row):
            valores[f"{idx}|{nome}"] = valor
    return valores

def _verificar_soma_total(df_sub, alvo):
    chaves = []
    soma   = 0.0
    for idx, row in df_sub.iterrows():
        ops = gerar_opcoes_nota(row)
        nome, valor = ops[-1]
        chaves.append(f"{idx}|{nome}")
        soma = round(soma + valor, 2)
    if abs(soma - alvo) <= TOLERANCIA:
        return chaves, soma
    return None, None

def conciliar(df_cli, valor_alvo, escopo="direto"):
    t0 = time.time()

    if escopo == "cidade":
        resultados_raw, meta, origem = _conciliar_cascata_cidade(df_cli, valor_alvo)
    else:
        chaves_total, soma_total = _verificar_soma_total(df_cli, valor_alvo)
        if chaves_total:
            resultados_raw = [(abs(soma_total - valor_alvo), chaves_total)]
            meta = {"pre": len(df_cli), "camada": "Soma Total"}
        else:
            valores = _montar_df_valores(df_cli)
            resultados_raw, meta = buscar_combinacoes(valores, valor_alvo, df_cli)
        origem = "cnpj"

    elapsed = round((time.time() - t0) * 1000)
    meta["tempo_ms"] = elapsed
    meta["origem"]   = origem

    resultados = []
    for dif, combo in resultados_raw:
        rbases, linhas = [], []
        total = 0
        for chave in combo:
            idx_str, tipo = chave.split("|")
            idx = int(idx_str)
            try:
                row = df_cli.loc[idx]
            except Exception:
                continue
            s   = round(float(row.get("saldo", 0)), 2)
            ir  = round(float(row.get("ir",    0)), 2)
            iss = round(float(row.get("iss",   0)), 2)

            if   tipo == "saldo":        val = s
            elif tipo == "saldo_ir":     val = round(s + ir,       2)
            elif tipo == "saldo_iss":    val = round(s + iss,      2)
            else:                        val = round(s + ir + iss, 2)

            retencao = round(s - val, 2)
            total    = round(total + val, 2)
            rb = str(row.get("rbase", ""))
            if rb and rb not in rbases: rbases.append(rb)

            linhas.append({
                "NF":            row.get("nf",    idx),
                "Nome":          str(row.get("nome",  "")),
                "CNPJ":          str(row.get("cnpj",  "")),
                "Venc.":         row.get("venc",  ""),
                "Atraso":        row.get("atraso", ""),
                "Saldo":         s,
                "IR":            ir,
                "ISS":           iss,
                "Vlr Utilizado": val,
                "Tipo":          tipo,
                "Retencao":      retencao,
            })

        resultados.append({
            "dif":   round(dif, 2),
            "total": round(total, 2),
            "rbase": " / ".join(rbases) or "N/A",
            "notas": pd.DataFrame(linhas),
        })

    return resultados, meta

# ============================================================
# BUSCA EM CASCATA PARA CIDADE
# ============================================================

def _conciliar_cascata_cidade(df_cidade, valor_alvo):
    col_cnpj = "cnpj_limpo" if "cnpj_limpo" in df_cidade.columns else "cnpj"
    for cnpj_val in df_cidade[col_cnpj].unique():
        df_sub = df_cidade[df_cidade[col_cnpj] == cnpj_val]
        nome_cnpj = df_sub["nome"].iloc[0] if "nome" in df_sub.columns else cnpj_val
        chaves_t, soma_t = _verificar_soma_total(df_sub, valor_alvo)
        if chaves_t:
            meta = {"pre": len(df_sub), "camada": "Soma Total",
                    "escopo_descricao": f"CNPJ individual: {nome_cnpj}"}
            return [(abs(soma_t - valor_alvo), chaves_t)], meta, "cnpj_individual"
        valores = _montar_df_valores(df_sub)
        res, meta = buscar_combinacoes(valores, valor_alvo, df_sub)
        if res:
            meta["escopo_descricao"] = f"CNPJ individual: {nome_cnpj}"
            return res, meta, "cnpj_individual"

    if "rbase" in df_cidade.columns:
        for rbase_val in df_cidade["rbase"].unique():
            df_sub = df_cidade[df_cidade["rbase"] == rbase_val]
            if df_sub["cnpj_limpo" if "cnpj_limpo" in df_cidade.columns else "cnpj"].nunique() < 2:
                continue
            chaves_t, soma_t = _verificar_soma_total(df_sub, valor_alvo)
            if chaves_t:
                meta = {"pre": len(df_sub), "camada": "Soma Total",
                        "escopo_descricao": f"Grupo RBASE {rbase_val} ({len(df_sub)} notas)"}
                return [(abs(soma_t - valor_alvo), chaves_t)], meta, "grupo_rbase"
            valores = _montar_df_valores(df_sub)
            res, meta = buscar_combinacoes(valores, valor_alvo, df_sub)
            if res:
                n_cnpjs = df_sub["cnpj"].nunique() if "cnpj" in df_sub.columns else "?"
                meta["escopo_descricao"] = f"Grupo RBASE {rbase_val} ({len(df_sub)} notas de {n_cnpjs} CNPJs)"
                return res, meta, "grupo_rbase"

    chaves_t, soma_t = _verificar_soma_total(df_cidade, valor_alvo)
    if chaves_t:
        meta = {"pre": len(df_cidade), "camada": "Soma Total",
                "escopo_descricao": f"Cidade completa ({df_cidade[col_cnpj].nunique()} CNPJs distintos)"}
        return [(abs(soma_t - valor_alvo), chaves_t)], meta, "cidade_ampla"
    valores = _montar_df_valores(df_cidade)
    res, meta = buscar_combinacoes(valores, valor_alvo, df_cidade)
    meta["escopo_descricao"] = f"Cidade completa ({df_cidade[col_cnpj].nunique()} CNPJs distintos)"
    return res, meta, "cidade_ampla"

# ============================================================
# HELPERS DE EXIBIÇÃO
# ============================================================

TIPO_BADGE = {
    "saldo":        ("🟢 Saldo",           "green"),
    "saldo_ir":     ("🟡 Saldo − IR",      "yellow"),
    "saldo_iss":    ("🟡 Saldo − ISS",     "yellow"),
    "saldo_ir_iss": ("🔴 Saldo − IR/ISS",  "red"),
}

ORIGEM_CONFIG = {
    "cnpj_individual": ("🎯 CNPJ individual",    "blue",   "Composição de notas de um único pagador."),
    "grupo_rbase":     ("🔗 Grupo RBASE",         "purple", "Composição mistura CNPJs do mesmo grupo econômico (mesma RBASE)."),
    "cidade_ampla":    ("⚠️  Cidade completa",    "yellow", "Atenção: composição mistura CNPJs de pagadores distintos. Confirme se houve pagamento conjunto."),
    "cnpj":            ("🎯 Busca direta",         "blue",   ""),
}

def brl(v):
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "—"

def fmt_tipo(tipo):
    label, _ = TIPO_BADGE.get(tipo, (tipo, "gray"))
    return label

# ============================================================
# EXIBIÇÃO DE COMBO — VERSÃO ATUALIZADA COM DESTAQUE CÉLULA
# ============================================================

# Mapa: qual variante → quais colunas ficam verdes
TIPO_COLUNAS_VERDES = {
    "saldo":        {"Saldo"},
    "saldo_ir":     {"Saldo", "IR"},
    "saldo_iss":    {"Saldo", "ISS"},
    "saldo_ir_iss": {"Saldo", "IR", "ISS"},
}

def exibir_combo(combo, i, n_total):
    is_best      = (i == 0)
    tem_retencao = combo["notas"]["Retencao"].abs().gt(TOLERANCIA).any()
    df           = combo["notas"].copy()

    tags = []
    if is_best:      tags.append('<span class="badge badge-blue">★ MELHOR OPÇÃO</span>')
    if tem_retencao: tags.append('<span class="badge badge-red">⚠ RETENÇÃO DETECTADA</span>')

    borda     = "1.5px solid #3b82f6" if is_best else "1px solid #1e293b"
    sombra    = "box-shadow:0 0 0 1px #3b82f620, 0 4px 24px #3b82f610;" if is_best else ""
    bg_card   = "#1e3a5f" if is_best else "#0a0f1a"
    bg_num    = "#3b82f6" if is_best else "#334155"
    cor_total = "#4ade80" if combo["dif"] == 0 else "#f59e0b"
    cor_delta = "#4ade80" if combo["dif"] == 0 else "#f87171"
    tags_html = "".join(tags)
    html_card = (
        "<div style='border:" + borda + ";border-radius:12px;overflow:hidden;margin-bottom:4px;" + sombra + "'>"
        "<div style='background:" + bg_card + ";padding:14px 20px;border-bottom:1px solid #1e293b;"
        "display:flex;align-items:center;gap:12px;flex-wrap:wrap;'>"
        "<span style='background:" + bg_num + ";color:#fff;border-radius:6px;padding:3px 12px;"
        "font-size:12px;font-weight:800;font-family:IBM Plex Mono,monospace;'>#" + str(i+1) + "</span>"
        + tags_html +
        "<span style='margin-left:auto;font-size:12px;color:#64748b;'>"
        + str(len(df)) + " nota(s) &nbsp;|&nbsp; RBASE: " + str(combo["rbase"])
        + " &nbsp;|&nbsp; Total: <b style='color:" + cor_total + ";'>" + brl(combo["total"]) + "</b>"
        + " &nbsp;|&nbsp; Δ: <b style='color:" + cor_delta + ";'>" + brl(combo["dif"]) + "</b>"
        + "</span></div></div>"
    )
    st.markdown(html_card, unsafe_allow_html=True)

    # ── Tabela com destaque célula-a-célula via HTML ──────────
    cnpjs_distintos = df["CNPJ"].nunique() if "CNPJ" in df.columns else 1
    colunas_exibir = ["NF", "Nome"]
    if cnpjs_distintos > 1:
        colunas_exibir.append("CNPJ")
    colunas_exibir += ["Venc.", "Atraso", "Saldo", "IR", "ISS", "Vlr Utilizado", "Tipo"]

    # Gerar tabela HTML com cores por célula
    html_rows = []
    for _, row in df.iterrows():
        tipo = row.get("Tipo", "saldo")
        colunas_verdes = TIPO_COLUNAS_VERDES.get(tipo, set())

        cells = []
        for col in colunas_exibir:
            val = row.get(col, "")
            # Formatar valores
            if col in ["Saldo", "IR", "ISS", "Vlr Utilizado"]:
                display_val = brl(val)
            elif col == "Venc.":
                display_val = val.strftime("%d/%m/%Y") if hasattr(val, "strftime") else str(val) if val else "—"
            elif col == "Atraso":
                display_val = f"{int(val)} dias" if val and val != 0 else "—"
            elif col == "Tipo":
                display_val = fmt_tipo(val)
            else:
                display_val = str(val) if val else "—"

            # Decidir cor
            if col in colunas_verdes:
                bg_c = "#22c55e20"; fc_c = "#4ade80"; fw = "700"
            elif col == "Vlr Utilizado":
                bg_c = "#3b82f620"; fc_c = "#60a5fa"; fw = "700"
            elif col == "Tipo":
                _, badge_cor = TIPO_BADGE.get(tipo, (tipo, "gray"))
                if badge_cor == "green":
                    bg_c = "#22c55e15"; fc_c = "#4ade80"; fw = "600"
                elif badge_cor == "yellow":
                    bg_c = "#f59e0b15"; fc_c = "#fbbf24"; fw = "600"
                else:
                    bg_c = "#ef444415"; fc_c = "#f87171"; fw = "600"
            else:
                bg_c = "transparent"; fc_c = "#94a3b8"; fw = "400"

            align = "right" if col in ["Saldo", "IR", "ISS", "Vlr Utilizado", "Atraso"] else "left"
            cells.append(
                f'<td style="padding:6px 10px;background:{bg_c};color:{fc_c};'
                f'font-weight:{fw};text-align:{align};font-size:12px;'
                f'border-bottom:1px solid #1e293b;white-space:nowrap;">{display_val}</td>'
            )
        html_rows.append("<tr>" + "".join(cells) + "</tr>")

    # Header
    header_cells = "".join(
        f'<th style="padding:6px 10px;background:#0f172a;color:#64748b;font-size:10px;'
        f'text-transform:uppercase;letter-spacing:0.05em;border-bottom:1px solid #334155;'
        f'text-align:left;white-space:nowrap;">{col}</th>'
        for col in colunas_exibir
    )

    html_table = (
        '<div style="border:1px solid #1e293b;border-radius:0 0 8px 8px;overflow-x:auto;margin-bottom:16px;">'
        '<table style="width:100%;border-collapse:collapse;font-family:IBM Plex Mono,monospace;">'
        f'<thead><tr>{header_cells}</tr></thead>'
        f'<tbody>{"".join(html_rows)}</tbody>'
        '</table></div>'
    )
    st.markdown(html_table, unsafe_allow_html=True)

    if tem_retencao:
        retidos = combo["notas"][combo["notas"]["Retencao"].abs() > TOLERANCIA]
        itens = []
        for _, r in retidos.iterrows():
            alerta = "IR+ISS" if r["IR"] != 0 and r["ISS"] != 0 else "IR" if r["IR"] != 0 else "ISS"
            nf_val = r["NF"]
            retencao_val = r['Retencao']
            itens.append(f"NF {nf_val} — retido <b>{brl(retencao_val)}</b> ({alerta})")
        br_itens = "<br>".join(itens)
        html_alerta = (
            '<div class="alert-retencao">' +
            '<b>\u26a0 Possível retenção indevida</b> — cliente deduziu IR/ISS do pagamento.<br>' +
            'Verifique isenção (ex: Simples Nacional) e solicite reembolso se aplicável.<br><br>' +
            br_itens + '</div>'
        )
        st.markdown(html_alerta, unsafe_allow_html=True)


# ============================================================
# CALCULADORA DE COMBINAÇÕES LIVRES (aba 2)
# ============================================================

def _parse_num(tok):
    import re
    tok = tok.strip().replace("\xa0", "").replace(" ", "")
    tok = re.sub(r"[R$]", "", tok)
    if tok in ("-", "–", "—", ""):
        return 0.0
    if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d+)?$", tok):
        tok = tok.replace(".", "").replace(",", ".")
    else:
        tok = tok.replace(",", ".")
    try:
        return float(tok)
    except ValueError:
        return None

def _parse_valores_livres(texto, modo="saldo"):
    import re
    rows = []
    for line in texto.splitlines():
        if "\t" in line:
            cols = re.split(r"\t", line)
        else:
            cols = re.split(r"[;|]", line)
        parsed = []
        for c in cols:
            v = _parse_num(c)
            if v is not None:
                parsed.append(v)
        if not parsed:
            continue
        if modo == "liquido":
            val = round(sum(parsed), 2)
        else:
            val = next((v for v in parsed if v > 0), None)
            if val is None:
                continue
            val = round(val, 2)
        if val > 0:
            rows.append(val)
    return rows

def _buscar_combinacoes_livres(valores_lista, alvo, max_n, top):
    valores = {str(i): v for i, v in enumerate(valores_lista)}
    valores = preprocessar(valores, alvo)
    if not valores:
        return []

    brutos = []
    g = greedy(valores, alvo, max_n)
    if g:
        brutos.append(g)
    mitm = meet_in_the_middle(valores, alvo, max_n)
    brutos.extend(mitm)

    if not brutos:
        brutos.extend(solver_milp(valores, alvo, max_n))

    vistas, resultado = set(), []
    for dif, combo in brutos:
        parcelas = tuple(sorted(round(float(valores.get(k, 0)), 2) for k in combo))
        if parcelas in vistas:
            continue
        vistas.add(parcelas)
        soma = round(sum(parcelas), 2)
        resultado.append((dif, soma, sorted(combo, key=lambda x: int(x))))

    resultado.sort(key=lambda x: (x[0], len(x[2])))
    return resultado[:top]

def aba_calculadora():
    st.markdown("""
    <div style="background:#0a1628; border:1px solid #1e293b; border-radius:12px;
         padding:20px 24px; margin-bottom:24px;">
        <div style="font-family:'IBM Plex Sans',sans-serif; font-weight:700;
             font-size:15px; color:#e2e8f0; margin-bottom:6px;">
            🧮 Calculadora de Combinações
        </div>
        <div style="font-size:12px; color:#64748b;">
            Cole uma lista de valores, informe o alvo e encontre quais combinações somam exatamente esse valor.
        </div>
    </div>
    """, unsafe_allow_html=True)

    c1, c2 = st.columns([2, 1])

    with c1:
        texto_vals = st.text_area(
            "Cole os valores aqui", height=220,
            placeholder="Formato simples (um por linha):\n124.170,17\n56.307,30\n47.724,75",
            key="calc_texto"
        )
        tem_tabs = "\t" in (texto_vals or "")
        if tem_tabs:
            modo = st.radio("Usar como valor:", ["Saldo bruto", "Líquido (Saldo + IR + ISS)"],
                           horizontal=True, key="calc_modo")
            modo_key = "liquido" if "Líquido" in modo else "saldo"
            if texto_vals.strip():
                preview = _parse_valores_livres(texto_vals, modo=modo_key)
                if preview:
                    st.caption(f"✅ {len(preview)} valores — soma: {brl(sum(preview))}")
        else:
            modo_key = "saldo"

    with c2:
        alvo_str = st.text_input("Valor alvo (R$)", placeholder="Ex: 436.611,53", key="calc_alvo")
        max_n = st.slider("Máx. parcelas", 1, 100, 8, key="calc_maxn")
        top_n = st.slider("Máx. resultados", 1, 20, 10, key="calc_top")
        st.markdown("<br>", unsafe_allow_html=True)
        calc_btn = st.button("🔍 Calcular", use_container_width=True, type="primary", key="calc_btn")

    if calc_btn:
        if not texto_vals.strip():
            st.error("Cole pelo menos um valor."); return
        if not alvo_str.strip():
            st.error("Informe o valor alvo."); return
        try:
            alvo = float(alvo_str.replace(".", "").replace(",", "."))
        except:
            st.error("Valor alvo inválido."); return

        vals = _parse_valores_livres(texto_vals, modo=modo_key)
        if not vals:
            st.error("Nenhum valor numérico reconhecido."); return

        st.markdown("---")
        s1, s2, s3, s4 = st.columns(4)
        s1.markdown(f'<div class="stat-box"><div class="stat-label">Valores</div><div class="stat-value">{len(vals)}</div></div>', unsafe_allow_html=True)
        s2.markdown(f'<div class="stat-box"><div class="stat-label">Soma total</div><div class="stat-value" style="font-size:15px;">{brl(sum(vals))}</div></div>', unsafe_allow_html=True)
        s3.markdown(f'<div class="stat-box"><div class="stat-label">Alvo</div><div class="stat-value green" style="font-size:15px;">{brl(alvo)}</div></div>', unsafe_allow_html=True)
        cor_d = "green" if abs(sum(vals)-alvo)<0.01 else "yellow"
        s4.markdown(f'<div class="stat-box"><div class="stat-label">Diferença</div><div class="stat-value {cor_d}" style="font-size:15px;">{brl(round(sum(vals)-alvo,2))}</div></div>', unsafe_allow_html=True)

        with st.spinner("Buscando..."):
            t0 = time.time()
            resultados = _buscar_combinacoes_livres(vals, alvo, max_n, top_n)
            elapsed = round((time.time() - t0) * 1000)

        if not resultados:
            st.error("Nenhuma combinação encontrada."); return

        st.markdown(f'<div class="alert-multi">⚡ <b>{len(resultados)} combinação(ões)</b> em <b>{elapsed}ms</b></div>', unsafe_allow_html=True)

        for i, (dif, soma, indices) in enumerate(resultados, 1):
            parcelas = [vals[int(idx)] for idx in indices]
            is_best = (i == 1)
            brd = "1.5px solid #3b82f6" if is_best else "1px solid #1e293b"
            bg_h = "#1e3a5f" if is_best else "#0a0f1a"
            cor_s = "#4ade80" if dif == 0 else "#f59e0b"
            st.markdown(
                f"<div style='border:{brd};border-radius:12px;overflow:hidden;margin-bottom:12px;'>"
                f"<div style='background:{bg_h};padding:10px 16px;display:flex;align-items:center;gap:10px;'>"
                f"<b style='color:#fff;font-family:monospace;'>#{i}</b>"
                f"<span style='margin-left:auto;font-size:12px;color:#94a3b8;font-family:monospace;'>"
                f"{len(parcelas)} parcela(s) | Soma: <b style='color:{cor_s};'>{brl(soma)}</b></span>"
                f"</div></div>", unsafe_allow_html=True
            )
            df_parc = pd.DataFrame({"Nº": [int(idx)+1 for idx in indices]+["TOTAL"], "Valor": [brl(v) for v in parcelas]+[brl(soma)]})
            st.dataframe(df_parc, use_container_width=True, hide_index=True)


# ============================================================
# ABA 3 — LIQUIDAÇÃO POR RETENÇÃO
# ============================================================

def aba_retencao():
    st.markdown("""
    <div style="background:#0a1628; border:1px solid #1e293b; border-radius:12px;
         padding:20px 24px; margin-bottom:24px;">
        <div style="font-weight:700; font-size:15px; color:#e2e8f0; margin-bottom:6px;">🎯 Notas Candidatas a Liquidação</div>
        <div style="font-size:12px; color:#64748b;">Identifica notas onde o saldo ≈ total retido (IR + ISS).</div>
    </div>
    """, unsafe_allow_html=True)

    up = st.file_uploader("📂 Planilha de títulos (.xlsx)", type=["xlsx","xls"], key="ret_upload")
    if not up:
        st.info("⬆️ Carregue a planilha para identificar candidatas."); return

    df = pd.read_excel(up)
    df.columns = df.columns.str.strip()
    MAPA_RET = {"NF":["NF","NR NFEM"],"NOME":["NOME CLIENTE","NOME"],"CIDADE":["CIDADE"],"CNPJ":["CNPJ"],
                "VLR SALDO":["SALDO","VLR SALDO"],"IR":["IR","VLR RETIDO"],"ISS":["ISS","ISS RETIDO"],
                "VENC":["VENCIMENTO","VENC","DT VENCIMENTO"],"ATRASO":["ATRASO"]}
    cols_up = {c.upper(): c for c in df.columns}
    for dest, cands in MAPA_RET.items():
        if dest not in df.columns:
            for cand in cands:
                if cand.upper() in cols_up:
                    df.rename(columns={cols_up[cand.upper()]: dest}, inplace=True); break
    for c in ["VLR SALDO","IR","ISS"]:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).round(2)
        else: df[c] = 0.0

    df = df[df["VLR SALDO"] > 0].copy()
    df["RETENCAO"] = (df["IR"].abs() + df["ISS"].abs()).round(2)
    df = df[df["RETENCAO"] > 0].copy()
    df["DIFERENÇA"] = (df["VLR SALDO"] - df["RETENCAO"]).abs().round(2)

    tolerancia = st.slider("Tolerância (R$)", 0.0, 50.0, 10.0, 0.5, format="R$ %.2f", key="ret_tol")
    df_cand = df[df["DIFERENÇA"] <= tolerancia].sort_values("DIFERENÇA")

    st.markdown("---")
    m1,m2,m3,m4 = st.columns(4)
    m1.markdown(f'<div class="stat-box"><div class="stat-label">Candidatas</div><div class="stat-value green">{len(df_cand):,}</div></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="stat-box"><div class="stat-label">Saldo</div><div class="stat-value" style="font-size:13px;">{brl(df_cand["VLR SALDO"].sum())}</div></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="stat-box"><div class="stat-label">Retido</div><div class="stat-value" style="font-size:13px;">{brl(df_cand["RETENCAO"].sum())}</div></div>', unsafe_allow_html=True)
    m4.markdown(f'<div class="stat-box"><div class="stat-label">Ajuste</div><div class="stat-value yellow" style="font-size:13px;">{brl(df_cand["DIFERENÇA"].sum())}</div></div>', unsafe_allow_html=True)

    if df_cand.empty:
        st.warning(f"Nenhuma nota com diferença ≤ R$ {tolerancia:.2f}."); return

    COLS_EXIB = ["NF","NOME","CIDADE","VLR SALDO","IR","ISS","RETENCAO","DIFERENÇA","VENC","ATRASO"]
    cols_ok = [c for c in COLS_EXIB if c in df_cand.columns]
    df_exib = df_cand[cols_ok].copy()
    df_export = df_exib.copy()
    for col in ["VLR SALDO","IR","ISS","RETENCAO","DIFERENÇA"]:
        if col in df_exib.columns:
            df_exib[col] = df_exib[col].apply(lambda v: brl(v) if pd.notna(v) else "")
    st.dataframe(df_exib, use_container_width=True, hide_index=True, height=480)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Liquidação")
    buf.seek(0)
    st.download_button("⬇️ Exportar Excel", buf, "liquidacao_retencao.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)


# ============================================================
# ABA ITAÚ (mantida igual)
# ============================================================

def extrair_liquidacoes_pdf(file_bytes):
    try:
        import pdfplumber
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber", "--break-system-packages", "-q"])
        import pdfplumber

    COL_SEU_NUM_MIN=280; COL_SEU_NUM_MAX=360; COL_TIPO_MIN=195; COL_TIPO_MAX=240
    COL_OPERACAO_MIN=620; COL_OPERACAO_MAX=760; COL_VALOR_FINAL_MIN=760; COL_VALOR_FINAL_MAX=850
    COL_JUROS_VAL_MIN=695; HEADER_Y_MAX=150
    def parse_val(t): return float(t.replace('.','').replace(',','.'))

    registros = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for num_pag, page in enumerate(pdf.pages, 1):
            words = page.extract_words()
            linhas = {}
            for w in words:
                y = round(w['top']/2)*2
                linhas.setdefault(y,[]).append(w)
            ultimo = None
            for y in sorted(linhas.keys()):
                if y < HEADER_Y_MAX: continue
                palavras = linhas[y]
                textos = [w['text'] for w in palavras]
                if 'juros' in textos and ultimo is not None:
                    for w in palavras:
                        if w['x0'] > COL_JUROS_VAL_MIN and w['text'] != 'juros':
                            try: ultimo['juros'] += parse_val(w['text'])
                            except: pass
                    continue
                seu_num=operacao=valor_final=None; pagador_words=[]
                for w in palavras:
                    x,t = w['x0'],w['text']
                    if COL_SEU_NUM_MIN<=x<=COL_SEU_NUM_MAX and _re.match(r'^\d{9,10}$',t): seu_num=t
                    if COL_OPERACAO_MIN<=x<=COL_OPERACAO_MAX and t not in ('----',): operacao=t
                    if x>=COL_VALOR_FINAL_MIN and _re.match(r'^[\d.,]+$',t) and t!='----':
                        try: valor_final=parse_val(t)
                        except: pass
                    if 45<=x<=195: pagador_words.append(t)
                if seu_num and operacao=='liquidação' and valor_final is not None:
                    reg = {'seu_num':seu_num,'pagador':' '.join(pagador_words),'valor_final':valor_final,'juros':0.0,'pagina':num_pag}
                    for w in palavras:
                        if 460<=w['x0']<=600 and _re.match(r'^[\d.,]+$',w['text']):
                            try: reg['valor_inicial']=parse_val(w['text'])
                            except: pass
                    if 'valor_inicial' not in reg: reg['valor_inicial']=None
                    registros.append(reg); ultimo=reg
                else:
                    if not (seu_num and operacao): ultimo=None
    return registros

def ler_csv_itau(file_bytes):
    import csv
    dados = {}
    text = file_bytes.decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(text), delimiter=';')
    for row in reader:
        doc = row.get('NUM_DOCUMENTO','').strip().strip('"')
        if not doc: continue
        def pv(k):
            try: return float(row.get(k,'0').strip().strip('"').replace('.','').replace(',','.'))
            except: return 0.0
        dados[doc] = {'nome':row.get('NOME','').strip().strip('"'),'parcela':pv('VLR_PARCELA_MOEDA_CORRENTE'),'recebido':pv('VLR_RECEB_MOEDA_CORRENTE')}
    return dados

def render_aba_itau():
    st.markdown("""
    <div style="background:#0a1628; border:1px solid #1e293b; border-radius:12px; padding:20px 24px; margin-bottom:24px;">
        <div style="font-weight:700; font-size:15px; color:#e2e8f0; margin-bottom:6px;">🏦 Conciliação Itaú — PDF × CSV</div>
        <div style="font-size:12px; color:#64748b;">Compara extrato Itaú (PDF) com recebimentos (CSV).</div>
    </div>
    """, unsafe_allow_html=True)
    col_pdf, col_csv = st.columns(2)
    with col_pdf: up_pdf = st.file_uploader("📄 Extrato (.pdf)", type=["pdf"], key="itau_pdf")
    with col_csv: up_csv = st.file_uploader("📋 Recebimentos (.csv)", type=["csv"], key="itau_csv")
    if not up_pdf or not up_csv:
        st.info("⬆️ Carregue PDF e CSV."); return

    with st.spinner("Lendo PDF..."):
        try: pdf_regs = extrair_liquidacoes_pdf(up_pdf.read())
        except Exception as e: st.error(f"Erro PDF: {e}"); return
    with st.spinner("Lendo CSV..."):
        try: csv_dados = ler_csv_itau(up_csv.read())
        except Exception as e: st.error(f"Erro CSV: {e}"); return
    if not pdf_regs:
        st.warning("Nenhuma liquidação no PDF."); return

    faltantes = [r for r in pdf_regs if r['seu_num'] not in csv_dados]
    diferencas = []
    for r in pdf_regs:
        if r['seu_num'] in csv_dados:
            diff = round(r['valor_final']-csv_dados[r['seu_num']]['recebido'],2)
            if abs(diff) > 0.009:
                diferencas.append({**r,'csv_recebido':csv_dados[r['seu_num']]['recebido'],'diff':diff})

    total_pdf = round(sum(r['valor_final'] for r in pdf_regs),2)
    total_csv = round(sum(v['recebido'] for v in csv_dados.values()),2)
    total_falt = round(sum(r['valor_final'] for r in faltantes),2)
    gap = round(total_pdf-total_csv,2)

    st.markdown("---")
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.markdown(f'<div class="stat-box"><div class="stat-label">PDF</div><div class="stat-value">{len(pdf_regs)}</div></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="stat-box"><div class="stat-label">CSV</div><div class="stat-value">{len(csv_dados)}</div></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="stat-box"><div class="stat-label">Faltantes</div><div class="stat-value {"red" if faltantes else "green"}">{len(faltantes)}</div></div>', unsafe_allow_html=True)
    m4.markdown(f'<div class="stat-box"><div class="stat-label">Diferenças</div><div class="stat-value {"yellow" if diferencas else "green"}">{len(diferencas)}</div></div>', unsafe_allow_html=True)
    m5.markdown(f'<div class="stat-box"><div class="stat-label">GAP</div><div class="stat-value {"red" if abs(gap)>0.01 else "green"}" style="font-size:13px;">{brl(gap)}</div></div>', unsafe_allow_html=True)

    st.markdown("### 1 · Faltantes")
    if faltantes:
        df_f = pd.DataFrame([{"Seu Número":r['seu_num'],"Pagador":r['pagador'],"Vlr Final":brl(r['valor_final']),"Página":r['pagina']} for r in faltantes])
        st.dataframe(df_f, use_container_width=True, hide_index=True)
    else:
        st.success("✔ Nenhum faltante.")

    st.markdown("### 2 · Diferenças")
    if diferencas:
        df_d = pd.DataFrame([{"Seu Número":d['seu_num'],"CSV":brl(d['csv_recebido']),"PDF":brl(d['valor_final']),"Diff":brl(d['diff'])} for d in diferencas])
        st.dataframe(df_d, use_container_width=True, hide_index=True)
    else:
        st.success("✔ Nenhuma diferença.")


# ============================================================
# ABA BB (mantida igual)
# ============================================================

_STOP_MATCH = {"MUN","MUNIC","MUNICIPAL","MUNICIPIO","PREFEITURA","PREF","FUNDO","SAUDE","SECRETARIA","CAMARA","SERVICO","SAMAE","FMS","SME","PM","CME","DE","DO","DA","DOS","DAS"}
_AMBIGUOS = {"SAUDE","ITA","CUSTODIA","BELA","VERDE","ALEGRE","NOVA","ALTO","ALTA","BOA","SOL","MAR"}

def _bb_normalize(text):
    if not text: return ""
    t = str(text).upper().strip()
    t = _unicodedata.normalize("NFD", t)
    return "".join(c for c in t if _unicodedata.category(c) != "Mn")

def _bb_extract_cnpj(text):
    if not text: return None
    s = str(text)
    m = _re.search(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}", s)
    if m: return _re.sub(r"\D", "", m.group())
    for token in s.split():
        clean = _re.sub(r"[.\-/]", "", token)
        if _re.match(r"^\d{14}$", clean): return clean
    return None

def _bb_consultar_cnpj(cnpj, cache):
    if cnpj in cache: return cache[cnpj]
    try:
        import requests
        r = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj}", timeout=8)
        result = (r.json().get("municipio","").upper(), r.json().get("uf","").upper()) if r.status_code == 200 else ("","")
    except: result = ("","")
    cache[cnpj] = result
    time.sleep(0.3)
    return result

@st.cache_data(show_spinner=False)
def _bb_carregar_agencias(file_bytes, file_name):
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="AGENCIAS")
    agencia_map, municipio_list, seen = {}, [], set()
    for _, row in df.iterrows():
        if not pd.notna(row.get("AGENCIA")): continue
        try: ag = str(int(float(str(row["AGENCIA"])))).zfill(4)
        except: ag = str(row["AGENCIA"]).strip().zfill(4)
        mun = str(row["MUNICIPIO"]).strip().upper() if pd.notna(row.get("MUNICIPIO")) else ""
        uf = str(row["UF"]).strip().upper() if pd.notna(row.get("UF")) else ""
        if ag and mun: agencia_map[ag] = (mun, uf)
        if mun:
            norm = _bb_normalize(mun)
            if norm not in seen: seen.add(norm); municipio_list.append((norm, mun, uf))
    municipio_list.sort(key=lambda x: -len(x[0]))
    return agencia_map, municipio_list

def _bb_buscar_municipio(texto, municipio_list):
    if not texto: return "", ""
    norm = _bb_normalize(texto)
    norm = _re.sub(r"^\d{2}/\d{2}\s+\d{2}:\d{2}\s+", "", norm)
    tokens_texto = set(_re.findall(r"[A-Z]{3,}", norm)) - _STOP_MATCH
    palavras_texto = [p for p in _re.findall(r"[A-Z]{4,}", norm) if p not in _STOP_MATCH]
    def word_match(nome, txt):
        return bool(_re.search(r"(?<![A-Z])"+_re.escape(nome)+r"(?![A-Z])", txt))
    for norm_mun, mun, uf in municipio_list:
        if len(norm_mun) < 5 or norm_mun in _AMBIGUOS: continue
        if word_match(norm_mun, norm): return mun, uf
    for norm_mun, mun, uf in municipio_list:
        tokens_mun = set(_re.findall(r"[A-Z]{3,}", norm_mun)) - _STOP_MATCH - _AMBIGUOS
        if tokens_mun and tokens_mun.issubset(tokens_texto): return mun, uf
    return "", ""

@st.cache_data(show_spinner=False)
def _bb_ler_extrato(file_bytes, file_name):
    import openpyxl as _opx
    wb = _opx.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    header_row, header_vals = None, []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if len([v for v in row if v is not None]) >= 3:
            header_row, header_vals = i, list(row); break
    if header_row is None: raise ValueError("Cabeçalho não encontrado.")
    rows_data = [list(row) for i, row in enumerate(ws.iter_rows(values_only=True), 1) if i > header_row and any(v is not None for v in row)]
    col_names, seen_cols = [], {}
    for v in header_vals:
        name = str(v).strip().upper() if v else "COL"
        if name in seen_cols: seen_cols[name] += 1; name = f"{name}_{seen_cols[name]}"
        else: seen_cols[name] = 0
        col_names.append(name)
    df = pd.DataFrame(rows_data, columns=col_names)
    df = df[df.apply(lambda r: any(str(v).strip() not in ("","None","nan","NAN") for v in r), axis=1)].copy()
    df.index = range(len(df))
    col_ag = next((c for c in df.columns if "AGENCI" in c), None)
    col_det = next((c for c in df.columns if "DETAL" in c), None) or next((c for c in df.columns if "HISTOR" in c), None)
    def _norm_ag(v):
        if v is None or str(v).strip() in ("","nan","None"): return "0000"
        try: return str(int(float(str(v).strip()))).zfill(4)
        except: return str(v).strip().zfill(4)
    df["_AGENCIA_ORIG"] = df[col_ag].apply(_norm_ag) if col_ag else "0000"
    df["_DETALHAMENTO"] = df[col_det].fillna("").astype(str).str.strip().str.upper() if col_det else ""
    return df

def render_aba_bb():
    st.markdown("""
    <div style="background:#0a1628; border:1px solid #1e293b; border-radius:12px; padding:20px 24px; margin-bottom:24px;">
        <div style="font-weight:700; font-size:15px; color:#e2e8f0; margin-bottom:6px;">🏛️ Conciliação BB — Extrato de Cobrança</div>
        <div style="font-size:12px; color:#64748b;">Enriquece extrato BB com UF e Município.</div>
    </div>
    """, unsafe_allow_html=True)
    col_ext, col_ag = st.columns(2)
    with col_ext: up_ext = st.file_uploader("📄 Extrato BB (.xlsx)", type=["xlsx","xls"], key="bb_extrato")
    with col_ag: up_ag = st.file_uploader("📋 Agências (.xlsx)", type=["xlsx","xls"], key="bb_agencias")
    if not up_ext or not up_ag:
        st.info("⬆️ Carregue extrato e agências."); return

    with st.spinner("Carregando agências..."):
        agencia_map, municipio_list = _bb_carregar_agencias(up_ag.read(), up_ag.name)
    with st.spinner("Lendo extrato..."):
        df = _bb_ler_extrato(up_ext.read(), up_ext.name)

    st.markdown("---")
    m1,m2,m3 = st.columns(3)
    m1.markdown(f'<div class="stat-box"><div class="stat-label">Agências</div><div class="stat-value">{len(agencia_map)}</div></div>', unsafe_allow_html=True)
    m2.markdown(f'<div class="stat-box"><div class="stat-label">Municípios</div><div class="stat-value">{len(municipio_list)}</div></div>', unsafe_allow_html=True)
    m3.markdown(f'<div class="stat-box"><div class="stat-label">Lançamentos</div><div class="stat-value green">{len(df)}</div></div>', unsafe_allow_html=True)

    if not st.button("🔍 Processar", type="primary", use_container_width=True, key="bb_proc"): return

    df["UF"]=""; df["MUNICIPIO"]=""; df["TITULO"]=""; df["STATUS"]=""
    cnpj_cache={}; n_cnpj=n_texto=n_agencia=n_verificar=0
    barra = st.progress(0, "Processando...")
    for i, (idx, row) in enumerate(df.iterrows()):
        det, ag = row["_DETALHAMENTO"], row["_AGENCIA_ORIG"]
        mun, uf = _bb_buscar_municipio(det, municipio_list)
        if mun: n_texto += 1
        if not mun:
            cnpj = _bb_extract_cnpj(det)
            if cnpj:
                mun, uf = _bb_consultar_cnpj(cnpj, cnpj_cache)
                if mun: n_cnpj += 1
        if not mun:
            mun, uf = agencia_map.get(ag, ("",""))
            if mun: n_agencia += 1
        if not mun: mun = "VERIFICAR"; n_verificar += 1
        df.at[idx,"UF"]=uf; df.at[idx,"MUNICIPIO"]=mun
        barra.progress(int((i+1)/len(df)*100), f"{i+1}/{len(df)}")
    barra.empty()

    st.markdown("---")
    r1,r2,r3,r4 = st.columns(4)
    r1.markdown(f'<div class="stat-box"><div class="stat-label">CNPJ</div><div class="stat-value green">{n_cnpj}</div></div>', unsafe_allow_html=True)
    r2.markdown(f'<div class="stat-box"><div class="stat-label">Texto</div><div class="stat-value green">{n_texto}</div></div>', unsafe_allow_html=True)
    r3.markdown(f'<div class="stat-box"><div class="stat-label">Agência</div><div class="stat-value">{n_agencia}</div></div>', unsafe_allow_html=True)
    r4.markdown(f'<div class="stat-box"><div class="stat-label">Verificar</div><div class="stat-value {"red" if n_verificar else "green"}">{n_verificar}</div></div>', unsafe_allow_html=True)

    orig_cols = [c for c in df.columns if not c.startswith("_") and c not in ("UF","MUNICIPIO","TITULO","STATUS")]
    st.dataframe(df[(orig_cols[:5]+["UF","MUNICIPIO"])].head(100), use_container_width=True, hide_index=True, height=400)


# ============================================================
# ABA 6 — CONCILIAÇÃO EM MASSA BB (NOVO — motor v5)
# ============================================================

def _massa_norm(texto):
    if pd.isna(texto): return ""
    t = str(texto).upper().strip()
    for p in ["PREFEITURA MUNICIPAL DE ","PREFEITURA MUNICIPAL ","FUNDO MUNICIPAL DE SAUDE DE ",
              "FUNDO MUNICIPAL DE SAUDE ","FUNDO DE SAUDE DE ","FUNDO DE SAUDE ",
              "MUNICIPIO DE ","SECRETARIA DE ","SECRETARIA MUNICIPAL DE ","SEC ","PM ","PREF ","PREF. ","FMS ","FME "]:
        t = t.replace(p, " ")
    return _re.sub(r'\s+', ' ', t).strip()

def _massa_sim(a, b):
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio()

def _massa_centavos(v):
    try: return round(float(v) * 100)
    except: return 0

def _massa_gerar_opcoes(row):
    """Gera variantes de valor para títulos (colunas SALDO, IR, ISS)."""
    s   = round(float(row.get("SALDO", 0) or 0), 2)
    ir  = round(float(row.get("IR", 0) or 0), 2)
    iss = round(float(row.get("ISS", 0) or 0), 2)
    ops = [("saldo", _massa_centavos(s))]
    if ir != 0:
        ops.append(("saldo_ir", _massa_centavos(round(s + ir, 2))))
    if iss != 0:
        ops.append(("saldo_iss", _massa_centavos(round(s + iss, 2))))
    if ir != 0 and iss != 0:
        ops.append(("liquido", _massa_centavos(round(s + ir + iss, 2))))
    # Deduplicar
    seen, deduped = set(), []
    for nome, val in ops:
        if val > 0 and val not in seen:
            seen.add(val); deduped.append((nome, val))
    return deduped

class _MassaTimeout(Exception): pass

def _massa_subset_multival(itens, alvo_cent, deadline, max_notas=30):
    """Solver multi-valor: acha subconjunto de itens cuja soma = alvo, máx 1 variante por título."""
    if not itens or alvo_cent <= 0: return "nenhum", []
    por_titulo = {}
    for idx, var, val in itens:
        if 0 < val <= alvo_cent:
            por_titulo.setdefault(idx, []).append((var, val))
    titulos = sorted(por_titulo.keys(), key=lambda t: max(v for _, v in por_titulo[t]), reverse=True)
    n = len(titulos)
    if n == 0: return "nenhum", []
    max_vals = [max(v for _, v in por_titulo[t]) for t in titulos]
    suf = [0]*(n+1)
    for i in range(n-1, -1, -1): suf[i] = suf[i+1]+max_vals[i]
    solucoes = []
    def bt(pos, rest, path, nu):
        if time.time() > deadline: raise _MassaTimeout()
        if rest == 0: solucoes.append(list(path)); return len(solucoes) < 2
        if pos >= n or nu >= max_notas or suf[pos] < rest: return True
        if not bt(pos+1, rest, path, nu): return False
        for var, val in por_titulo[titulos[pos]]:
            if val > rest: continue
            path.append((titulos[pos], var, val))
            if not bt(pos+1, rest-val, path, nu+1): path.pop(); return False
            path.pop()
        return True
    try: bt(0, alvo_cent, [], 0)
    except _MassaTimeout:
        if len(solucoes)==1: return "unico", solucoes[0]
        return "timeout", []
    if len(solucoes)==0: return "nenhum", []
    if len(solucoes)==1: return "unico", solucoes[0]
    return "ambiguo", []

def _massa_tentar_match(itens, alvo_cent, deadline):
    status, sol = _massa_subset_multival(itens, alvo_cent, deadline)
    if status == "unico": return "unico", sol
    if status == "nenhum":
        tol = max(1, round(alvo_cent * 0.02))
        for delta in range(1, tol+1):
            for sinal in (1, -1):
                apr = alvo_cent + sinal*delta
                if apr <= 0: continue
                s2, sl2 = _massa_subset_multival(itens, apr, min(deadline, time.time()+2))
                if s2 == "unico": return "unico_aprox", sl2
    return status, sol

def _massa_selecionar(pag, titulos):
    mun = pag.get("municipio_norm", "")
    if mun:
        exatos = titulos[titulos["cidade_norm"] == mun]
        if len(exatos) > 0:
            return exatos, f"Município Exato ({pag.get('municipio','')})"
    if mun and len(mun) >= 3:
        scores = titulos["cidade_norm"].apply(lambda c: _massa_sim(mun, c) if c else 0)
        mask = scores >= 0.70
        if mask.any():
            best = scores[mask].max()
            cidade = titulos.loc[scores==best, "CIDADE"].iloc[0]
            return titulos[titulos["CIDADE"]==cidade], f"Município Fuzzy ({cidade}, {best:.0%})"
    hist = pag.get("historico_norm", "")
    if hist and len(hist) >= 5:
        scores = titulos["nome_norm"].apply(lambda n: _massa_sim(hist, n) if n else 0)
        mask = scores >= 0.65
        if mask.any():
            best = scores[mask].max()
            cidade = titulos.loc[scores==best, "CIDADE"].iloc[0]
            return titulos[titulos["CIDADE"]==cidade], f"Histórico Fuzzy ({best:.0%})"
    return pd.DataFrame(), "Sem candidatos"


VARIANTE_LABEL_MASSA = {"saldo":"Saldo","saldo_ir":"Saldo−IR","saldo_iss":"Saldo−ISS","liquido":"Líquido"}

def render_aba_massa_bb():
    st.markdown("""
    <div style="background:#0a1628; border:1px solid #1e293b; border-radius:12px; padding:20px 24px; margin-bottom:24px;">
        <div style="font-weight:700; font-size:15px; color:#e2e8f0; margin-bottom:6px;">
            📊 Conciliação em Massa BB — Motor v5
        </div>
        <div style="font-size:12px; color:#64748b;">
            Cruza TODOS os pagamentos da planilha de pendências com os títulos em aberto.
            Usa solver multi-valor (Saldo, Saldo−IR, Saldo−ISS, Líquido) para encontrar
            a composição exata de notas. Gera Excel com resultado completo.
        </div>
    </div>
    """, unsafe_allow_html=True)

    col_pend, col_tit = st.columns(2)
    with col_pend:
        up_pend = st.file_uploader("📂 Planilha de Pendências (.xlsx)", type=["xlsx","xls"], key="massa_pend",
                                    help="Planilha com os pagamentos a conciliar (colunas: DT LANÇAMENTO, VALOR, HISTÓRICO, MUNICÍPIO, STATUS)")
    with col_tit:
        up_tit = st.file_uploader("📂 Títulos em Aberto (.xlsx)", type=["xlsx","xls"], key="massa_tit",
                                   help="Planilha com os títulos/NFs (colunas: SALDO, IR, ISS, CIDADE, NOME CLIENTE)")

    if not up_pend or not up_tit:
        st.info("⬆️ Carregue as duas planilhas para iniciar a conciliação em massa.")
        st.markdown("""
**Como funciona:**
1. Carregue a planilha de **pendências** (pagamentos do banco) e a de **títulos em aberto**
2. Selecione a **aba** da planilha de pendências a analisar
3. O motor cruza cada pagamento com os títulos, testando **4 variantes de valor** por nota
4. Resultado: Excel com blocos verde (composição encontrada) e aba de não resolvidos

**O solver multi-valor testa para cada nota:**
- 🟢 **Saldo** — valor integral
- 🟡 **Saldo − IR** — pagador deduziu IR
- 🟡 **Saldo − ISS** — pagador deduziu ISS
- 🔴 **Líquido** — pagador deduziu IR + ISS
        """)
        return

    # Ler abas disponíveis
    try:
        xl = pd.ExcelFile(io.BytesIO(up_pend.read()))
        up_pend.seek(0)
        abas_ignorar = {"classificação", "PORTAIS", "PG INTERMUNICIPAIS", "ADIANTAMENTOS"}
        abas = [s for s in xl.sheet_names if s not in abas_ignorar]
    except Exception as e:
        st.error(f"Erro ao ler planilha: {e}"); return

    if not abas:
        st.error("Nenhuma aba encontrada."); return

    c1, c2 = st.columns([2, 1])
    with c1:
        aba_sel = st.selectbox("Selecione a aba de pendências", abas, key="massa_aba")
    with c2:
        st.markdown("<br>", unsafe_allow_html=True)
        executar = st.button("🚀 Executar Conciliação em Massa", type="primary", use_container_width=True, key="massa_exec")

    if not executar:
        return

    # ── Ler pendências ────────────────────────────────────
    with st.spinner(f"Lendo aba '{aba_sel}'..."):
        try:
            pend_bytes = up_pend.read()
            up_pend.seek(0)
            raw = pd.read_excel(io.BytesIO(pend_bytes), sheet_name=aba_sel, dtype=str, header=None)
            header_row = None
            for i, row in raw.iterrows():
                vals = [str(v).strip().upper() for v in row if not pd.isna(v)]
                if any(v in ("VALOR","DT LANÇAMENTO","HISTORICO","HISTÓRICO") for v in vals):
                    header_row = i; break
            if header_row is None:
                st.error("Cabeçalho não encontrado na aba."); return
            df_raw = raw.iloc[header_row:].copy()
            df_raw.columns = df_raw.iloc[0].str.strip()
            df_raw = df_raw.iloc[1:].reset_index(drop=True)

            col_map = {"data":["DT LANÇAMENTO","DATA"],"banco":["BANCO"],"status":["STATUS"],
                       "valor":["VALOR","VALOR PAGO","VL PAGO"],"historico":["HISTÓRICO","HISTORICO"],
                       "municipio":["MUNICIPIO","MUNICÍPIO","CIDADE"],"obs":["OBSERVAÇAO","OBSERVAÇÃO","OBS"]}
            cols_disp = {c.strip().upper(): c for c in df_raw.columns if not pd.isna(c)}
            mapa = {}
            for campo, opcoes in col_map.items():
                mapa[campo] = next((cols_disp[o.upper()] for o in opcoes if o.upper() in cols_disp), None)

            banco = pd.DataFrame()
            banco["data"] = pd.to_datetime(df_raw[mapa["data"]], errors="coerce") if mapa["data"] else pd.NaT
            banco["banco"] = df_raw[mapa["banco"]].str.strip() if mapa["banco"] else ""
            banco["status"] = df_raw[mapa["status"]].str.strip() if mapa["status"] else ""
            banco["valor"] = pd.to_numeric(df_raw[mapa["valor"]].astype(str).str.replace(",","."), errors="coerce") if mapa["valor"] else 0
            banco["historico"] = df_raw[mapa["historico"]].str.strip() if mapa["historico"] else ""
            banco["municipio"] = df_raw[mapa["municipio"]].str.strip().str.upper() if mapa["municipio"] else ""
            banco["obs"] = df_raw[mapa["obs"]].str.strip() if mapa["obs"] else ""
            banco["municipio_norm"] = banco["municipio"].apply(_massa_norm)
            banco["historico_norm"] = banco["historico"].apply(_massa_norm)
            banco["valor_cent"] = banco["valor"].apply(_massa_centavos)
            banco = banco.dropna(subset=["valor"])
            banco = banco[banco["valor"] > 0]
            # Filtrar BAIXADOS
            mask_b = banco["status"].str.upper().str.strip().isin({"BAIXADO","BAIXADO ","baixado"})
            n_baixados = mask_b.sum()
            banco = banco[~mask_b].reset_index(drop=True)
        except Exception as e:
            st.error(f"Erro ao ler pendências: {e}"); return

    # ── Ler títulos ───────────────────────────────────────
    with st.spinner("Lendo títulos..."):
        try:
            tit_bytes = up_tit.read()
            titulos = pd.read_excel(io.BytesIO(tit_bytes), dtype=str)
            titulos.columns = titulos.columns.str.strip()
            for col in ["SALDO","BRUTO","CORRETO","IR","ISS"]:
                if col in titulos.columns:
                    titulos[col] = pd.to_numeric(titulos[col].astype(str).str.replace(",","."), errors="coerce")
            for col in ["SALDO","IR","ISS"]:
                if col not in titulos.columns: titulos[col] = 0.0
                titulos[col] = titulos[col].fillna(0.0)
            if "CORRETO" not in titulos.columns:
                titulos["CORRETO"] = (titulos["SALDO"]+titulos["IR"]+titulos["ISS"]).round(2)
            titulos["VENCIMENTO"] = pd.to_datetime(titulos.get("VENCIMENTO", pd.NaT), errors="coerce")
            titulos["CIDADE"] = titulos["CIDADE"].str.strip().str.upper() if "CIDADE" in titulos.columns else ""
            titulos["UF"] = titulos["UF"].str.strip().str.upper() if "UF" in titulos.columns else ""
            titulos["cidade_norm"] = titulos["CIDADE"].apply(_massa_norm)
            titulos["nome_norm"] = titulos["NOME CLIENTE"].apply(_massa_norm) if "NOME CLIENTE" in titulos.columns else ""
            titulos["idx_titulo"] = range(len(titulos))
            titulos = titulos[(titulos["SALDO"].notna()) & (titulos["SALDO"] > 0)].copy()
        except Exception as e:
            st.error(f"Erro ao ler títulos: {e}"); return

    # ── Stats ─────────────────────────────────────────────
    st.markdown("---")
    s1,s2,s3,s4 = st.columns(4)
    s1.markdown(f'<div class="stat-box"><div class="stat-label">Pagamentos</div><div class="stat-value">{len(banco)}</div></div>', unsafe_allow_html=True)
    s2.markdown(f'<div class="stat-box"><div class="stat-label">Baixados (ignorados)</div><div class="stat-value">{n_baixados}</div></div>', unsafe_allow_html=True)
    s3.markdown(f'<div class="stat-box"><div class="stat-label">Títulos</div><div class="stat-value green">{len(titulos)}</div></div>', unsafe_allow_html=True)
    s4.markdown(f'<div class="stat-box"><div class="stat-label">Total pendente</div><div class="stat-value" style="font-size:14px;">{brl(banco["valor"].sum())}</div></div>', unsafe_allow_html=True)

    # ── Executar conciliação ──────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)

    resultados = []
    titulos_usados = set()
    banco["chave_grupo"] = banco["municipio"].fillna("")+"|"+banco["data"].dt.strftime("%Y-%m-%d").fillna("")
    grupos = list(banco.groupby("chave_grupo", sort=False))

    barra = st.progress(0, "Conciliando pagamentos...")
    total_pags = len(banco)
    pag_processados = 0

    for g_idx, (chave, grupo) in enumerate(grupos):
        pags = grupo.sort_values("valor", ascending=False)
        ref = pags.iloc[0]
        cands, fase = _massa_selecionar(ref, titulos)

        if len(cands) == 0:
            for _, p in pags.iterrows():
                resultados.append({"pag":p,"status":"SEM_CANDIDATOS","fase":fase,"solucao":[],"obs":"Município não encontrado"})
                pag_processados += 1
            barra.progress(min(99, int(pag_processados/total_pags*100)), f"{pag_processados}/{total_pags}")
            continue

        cands_livres = cands[~cands["idx_titulo"].isin(titulos_usados)].copy()
        if len(cands_livres) == 0:
            for _, p in pags.iterrows():
                resultados.append({"pag":p,"status":"SEM_CANDIDATOS","fase":fase,"solucao":[],"obs":"Títulos já vinculados"})
                pag_processados += 1
            barra.progress(min(99, int(pag_processados/total_pags*100)), f"{pag_processados}/{total_pags}")
            continue

        # Montar itens multi-valor
        itens_base = []
        for _, row in cands_livres.iterrows():
            for var, val in _massa_gerar_opcoes(row):
                if val > 0: itens_base.append((int(row["idx_titulo"]), var, val))

        deadline = time.time() + 15

        # Cada pagamento individual
        usados_grupo = set()
        for _, p in pags.iterrows():
            alvo_cent = int(p["valor_cent"])
            itens_livres = [(i,v,vl) for i,v,vl in itens_base if i not in usados_grupo]
            if not itens_livres:
                resultados.append({"pag":p,"status":"SEM_CANDIDATOS","fase":fase,"solucao":[],"obs":"Esgotados"})
            else:
                dl = min(deadline, time.time()+8)
                status, sol = _massa_tentar_match(itens_livres, alvo_cent, dl)
                if status in ("unico","unico_aprox"):
                    obs = "Valor aprox. (≤2%)" if status=="unico_aprox" else ""
                    resultados.append({"pag":p,"status":status,"fase":fase,"solucao":sol,"obs":obs})
                    for idx_t,_,_ in sol:
                        usados_grupo.add(idx_t); titulos_usados.add(idx_t)
                else:
                    obs_map = {"ambiguo":"Múltiplas combinações","timeout":"Timeout","nenhum":"Sem combinação"}
                    resultados.append({"pag":p,"status":status,"fase":fase,"solucao":[],"obs":obs_map.get(status,status)})

            pag_processados += 1
            barra.progress(min(99, int(pag_processados/total_pags*100)), f"{pag_processados}/{total_pags}")

    barra.progress(100, "Concluído!")
    time.sleep(0.3)
    barra.empty()

    # ── Resultados ────────────────────────────────────────
    n_ok = sum(1 for r in resultados if r["status"] in ("unico","unico_aprox"))
    n_apr = sum(1 for r in resultados if r["status"]=="unico_aprox")
    n_amb = sum(1 for r in resultados if r["status"]=="ambiguo")
    n_sem = len(resultados) - n_ok - n_amb
    taxa = n_ok/len(resultados)*100 if resultados else 0

    st.markdown("---")
    r1,r2,r3,r4,r5 = st.columns(5)
    r1.markdown(f'<div class="stat-box"><div class="stat-label">Total</div><div class="stat-value">{len(resultados)}</div></div>', unsafe_allow_html=True)
    r2.markdown(f'<div class="stat-box"><div class="stat-label">✅ Conciliados</div><div class="stat-value green">{n_ok}</div></div>', unsafe_allow_html=True)
    r3.markdown(f'<div class="stat-box"><div class="stat-label">⚠️ Ambíguos</div><div class="stat-value yellow">{n_amb}</div></div>', unsafe_allow_html=True)
    r4.markdown(f'<div class="stat-box"><div class="stat-label">❌ Sem match</div><div class="stat-value red">{n_sem}</div></div>', unsafe_allow_html=True)
    cor_taxa = "green" if taxa >= 80 else "yellow" if taxa >= 50 else "red"
    r5.markdown(f'<div class="stat-box"><div class="stat-label">Taxa</div><div class="stat-value {cor_taxa}">{taxa:.0f}%</div></div>', unsafe_allow_html=True)

    if n_ok > 0:
        st.markdown(f'<div class="alert-success">✅ <b>{n_ok} pagamentos conciliados</b> ({n_ok-n_apr} exatos + {n_apr} aproximados)</div>', unsafe_allow_html=True)

    # ── Mostrar exemplos ──────────────────────────────────
    tit_map = {int(row["idx_titulo"]): row for _, row in titulos.iterrows()}

    ok_results = [r for r in resultados if r["status"] in ("unico","unico_aprox")]
    nok_results = [r for r in resultados if r["status"] not in ("unico","unico_aprox")]

    if ok_results:
        st.markdown("### ✅ Conciliados (primeiros 20)")
        for res in ok_results[:20]:
            pag = res["pag"]
            sol = res["solucao"]
            aprox = res["status"]=="unico_aprox"
            soma_comp = round(sum(v/100 for _,_,v in sol),2) if sol else 0

            badge_a = ' <span class="badge badge-yellow">APROX</span>' if aprox else ""
            st.markdown(
                f'<div style="background:#0a1628;border:1px solid #1e293b;border-radius:8px;padding:10px 16px;margin-bottom:4px;">'
                f'<span style="color:#64748b;font-size:11px;">{pag.get("data","").strftime("%d/%m") if hasattr(pag.get("data",""),"strftime") else ""}</span> '
                f'<b style="color:#e2e8f0;">{pag.get("municipio","")}</b> — '
                f'<span style="color:#4ade80;font-family:monospace;font-weight:700;">{brl(pag["valor"])}</span>'
                f'{badge_a} → '
                f'<span style="color:#94a3b8;font-size:12px;">{len(sol)} nota(s): '
                + ", ".join(f'NF {tit_map.get(i,{}).get("NF","?")} ({VARIANTE_LABEL_MASSA.get(v,v)} R${vl/100:,.2f})' for i,v,vl in sol)
                + f'</span></div>', unsafe_allow_html=True
            )

    if nok_results:
        st.markdown("### ❌ Não resolvidos (primeiros 20)")
        for res in nok_results[:20]:
            pag = res["pag"]
            st.markdown(
                f'<div style="background:#1a0a0a;border:1px solid #ef444430;border-radius:8px;padding:8px 16px;margin-bottom:4px;">'
                f'<span style="color:#64748b;font-size:11px;">{pag.get("data","").strftime("%d/%m") if hasattr(pag.get("data",""),"strftime") else ""}</span> '
                f'<b style="color:#f87171;">{pag.get("municipio","")}</b> — '
                f'<span style="color:#fca5a5;font-family:monospace;">{brl(pag["valor"])}</span> '
                f'<span style="color:#64748b;font-size:11px;">| {res["obs"]}</span></div>', unsafe_allow_html=True
            )

    # ── Gerar Excel para download ─────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Exportar Resultado Completo")

    with st.spinner("Gerando Excel..."):
        from openpyxl import Workbook
        from openpyxl.styles import Font as XlFont, PatternFill as XlFill, Alignment as XlAlign, Border as XlBorder, Side as XlSide

        wb = Workbook(); wb.remove(wb.active)
        FONTE = "Arial"
        def _borda():
            s = XlSide(style="thin", color="C0C0C0")
            return XlBorder(left=s, right=s, top=s, bottom=s)
        def _cel(ws, r, c, v=None, bold=False, color="000000", bg=None, sz=9, ha="left"):
            cell = ws.cell(row=r, column=c)
            if v is not None: cell.value = v
            cell.font = XlFont(name=FONTE, bold=bold, color=color, size=sz)
            cell.alignment = XlAlign(horizontal=ha, vertical="center")
            cell.border = _borda()
            if bg: cell.fill = XlFill("solid", start_color=bg)
            return cell

        # Aba CONCILIAÇÃO
        ws = wb.create_sheet("🗂️ CONCILIAÇÃO")
        ws.sheet_view.showGridLines = False
        COLS = [("NF",10),("DOC",14),("VENCIMENTO",13),("SALDO",15),("IR",11),("ISS",11),("CORRETO",15),("NOME CLIENTE",38),("UF",6),("TIPO",14)]
        for i,(_, w) in enumerate(COLS,1): ws.column_dimensions[chr(64+i) if i<=26 else "A"].width = w

        ws.merge_cells("A1:J1")
        ws["A1"].value = f"CONCILIAÇÃO EM MASSA — {aba_sel}"
        ws["A1"].font = XlFont(name=FONTE, bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = XlFill("solid", start_color="1F3864")
        ws["A1"].alignment = XlAlign(horizontal="center", vertical="center")

        VERDE_BG_X="C6EFCE"; VERDE_FT_X="276221"; NEUTRO_BG_X="F7F7F7"; NEUTRO_FT_X="666666"
        AMAR_BG_X="FFEB9C"; AMAR_FT_X="9C6500"
        CAMPOS_VERDES = {"saldo":{"SALDO"},"saldo_ir":{"SALDO","IR"},"saldo_iss":{"SALDO","ISS"},"liquido":{"SALDO","IR","ISS","CORRETO"}}

        row = 3
        for res in ok_results:
            pag = res["pag"]; sol = res["solucao"]
            aprox = res["status"]=="unico_aprox"
            # Header pagamento
            hdrs = ["DATA","BANCO","VALOR PAGO","HISTÓRICO","MUNICÍPIO","","","","",""]
            for c, h in enumerate(hdrs,1): _cel(ws,row,c,h,bold=True,color="FFFFFF",bg="1F3864",ha="center",sz=8)
            row += 1
            vals = [pag["data"].date() if pd.notna(pag.get("data")) else "",pag.get("banco",""),pag["valor"],
                    pag.get("historico",""),pag.get("municipio",""),"","","","",""]
            if aprox: vals[9] = "⚠ Aprox (≤2%)"
            for c, v in enumerate(vals,1):
                cell = _cel(ws,row,c,v,bold=True,color="FFFFFF",bg="2E4057",ha="center" if c<=3 else "left",sz=10)
                if c==3: cell.number_format="#,##0.00"
            if aprox: _cel(ws,row,10,"⚠ Aprox",bold=True,color=AMAR_FT_X,bg=AMAR_BG_X,ha="center",sz=8)
            row += 1
            # Header colunas
            for c,(campo,_) in enumerate(COLS,1): _cel(ws,row,c,campo,bold=True,color="FFFFFF",bg="2E75B6",ha="center",sz=8)
            row += 1
            # Notas da composição
            for idx_t, variante, val_cent in sol:
                tit = tit_map.get(idx_t)
                if not tit: continue
                cv = CAMPOS_VERDES.get(variante, set())
                for c,(campo,_) in enumerate(COLS,1):
                    if campo == "TIPO":
                        val = VARIANTE_LABEL_MASSA.get(variante, variante)
                        _cel(ws,row,c,val,bold=True,color=AMAR_FT_X,bg=AMAR_BG_X,ha="center")
                    else:
                        val = tit.get(campo,"")
                        if campo=="VENCIMENTO" and pd.notna(val):
                            try: val = val.date()
                            except: pass
                        if campo in cv:
                            cell = _cel(ws,row,c,val,bold=True,color=VERDE_FT_X,bg=VERDE_BG_X,ha="right" if campo in ("SALDO","IR","ISS","CORRETO") else "center")
                        else:
                            cell = _cel(ws,row,c,val,color=NEUTRO_FT_X,bg=NEUTRO_BG_X,ha="right" if campo in ("SALDO","IR","ISS","CORRETO") else "center")
                        if campo in ("SALDO","IR","ISS","CORRETO"): cell.number_format="#,##0.00"
                row += 1
            # Separador
            for c in range(1,11): ws.cell(row=row,column=c).fill = XlFill("solid", start_color="F0F0F0")
            row += 1

        # Aba NÃO RESOLVIDOS
        ws2 = wb.create_sheet("❌ NÃO RESOLVIDOS")
        hdrs2 = ["Data","Banco","Valor","Histórico","Município","Status","Fase","Motivo"]
        for c,h in enumerate(hdrs2,1): _cel(ws2,1,c,h,bold=True,color="FFFFFF",bg="9C0006",ha="center")
        for i,res in enumerate(nok_results,2):
            pag = res["pag"]
            status_lbl = {"ambiguo":"⚠️ Ambíguo","nenhum":"❌ Sem combinação","SEM_CANDIDATOS":"❌ Sem candidatos","timeout":"⏱️ Timeout"}.get(res["status"],res["status"])
            vals2 = [pag["data"].date() if pd.notna(pag.get("data")) else "",pag.get("banco",""),pag["valor"],
                     pag.get("historico",""),pag.get("municipio",""),status_lbl,res.get("fase",""),res.get("obs","")]
            for c,v in enumerate(vals2,1):
                cell = _cel(ws2,i,c,v,color="9C0006",bg="FFC7CE")
                if c==3: cell.number_format="#,##0.00"
        for l,w in zip("ABCDEFGH",[12,8,14,40,20,20,22,40]): ws2.column_dimensions[l].width = w

        # Aba RESUMO
        ws3 = wb.create_sheet("📊 RESUMO")
        ws3.merge_cells("A1:D1")
        ws3["A1"].value = f"RESUMO — {aba_sel}"; ws3["A1"].font = XlFont(name=FONTE,bold=True,size=13,color="FFFFFF")
        ws3["A1"].fill = XlFill("solid",start_color="1F3864"); ws3["A1"].alignment = XlAlign(horizontal="center")
        for c,h in enumerate(["Situação","Qtd","Valor (R$)","% Total"],1): _cel(ws3,3,c,h,bold=True,color="FFFFFF",bg="2E75B6",ha="center")
        v_tot = sum(r["pag"]["valor"] for r in resultados) or 1
        linhas = [
            ("Total",len(resultados),sum(r["pag"]["valor"] for r in resultados),1.0,None),
            ("✅ Exatos",n_ok-n_apr,sum(r["pag"]["valor"] for r in resultados if r["status"]=="unico"),None,"C6EFCE"),
            ("🟡 Aproximados",n_apr,sum(r["pag"]["valor"] for r in resultados if r["status"]=="unico_aprox"),None,"FFEB9C"),
            ("⚠️ Ambíguos",n_amb,sum(r["pag"]["valor"] for r in resultados if r["status"]=="ambiguo"),None,"FFEB9C"),
            ("❌ Sem match",n_sem,sum(r["pag"]["valor"] for r in resultados if r["status"] not in ("unico","unico_aprox","ambiguo")),None,"FFC7CE"),
        ]
        for i,(lbl,qtd,val,pct,bg) in enumerate(linhas,4):
            _cel(ws3,i,1,lbl,bg=bg); _cel(ws3,i,2,qtd,ha="center",bg=bg)
            cell = _cel(ws3,i,3,val,ha="right",bg=bg); cell.number_format="#,##0.00"
            if pct: cell2 = _cel(ws3,i,4,pct,ha="center",bg=bg); cell2.number_format="0.0%"
            else: _cel(ws3,i,4,round(val/v_tot,3) if v_tot else 0,ha="center",bg=bg).number_format="0.0%"
        for l,w in zip("ABCD",[40,10,18,12]): ws3.column_dimensions[l].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

    st.download_button(
        "⬇️ Baixar Excel Completo",
        buf,
        f"conciliacao_massa_{aba_sel.replace(' ','_')}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="massa_download"
    )


# ============================================================
# INTERFACE PRINCIPAL
# ============================================================

st.markdown("""
<div class="header-box">
    <div style="display:flex; align-items:center; gap:16px;">
        <div style="font-size:36px;">⚖️</div>
        <div>
            <div style="font-family:'IBM Plex Sans',sans-serif; font-weight:800;
                 font-size:22px; letter-spacing:-0.02em; color:#f1f5f9;">
                Motor de Conciliação Financeira
            </div>
            <div style="font-size:12px; color:#475569; letter-spacing:0.05em; margin-top:2px;">
                MAXIFROTA — ANÁLISE DE TÍTULOS VENCIDOS • MOTOR v5
            </div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

aba_conciliacao, aba_massa, aba_calc, aba_ret, aba_itau, aba_bb = st.tabs([
    "⚖️  Conciliação por Planilha",
    "📊  Conciliação em Massa BB",
    "🧮  Calculadora de Combinações",
    "🎯  Liquidação por Retenção",
    "🏦  Conciliação Itaú",
    "🏛️  Conciliação BB",
])

# ══════════════════════════════════════════════════════════════
# ABA 1 — CONCILIAÇÃO POR PLANILHA (busca individual)
# ══════════════════════════════════════════════════════════════
with aba_conciliacao:

    uploaded = st.file_uploader("📂 Carregar planilha de pendências", type=["xlsx", "xls"],
                                 help="Arraste o arquivo .xlsx com os títulos vencidos")

    if uploaded:
        with st.spinner("Lendo planilha..."):
            df = carregar_planilha(uploaded.read(), uploaded.name)

        col1, col2, col3 = st.columns(3)
        col1.markdown(f'<div class="stat-box"><div class="stat-label">Arquivo</div><div class="stat-value" style="font-size:14px;">{uploaded.name}</div></div>', unsafe_allow_html=True)
        col2.markdown(f'<div class="stat-box"><div class="stat-label">Títulos</div><div class="stat-value green">{len(df):,}</div></div>', unsafe_allow_html=True)
        col3.markdown(f'<div class="stat-box"><div class="stat-label">Colunas</div><div class="stat-value">{len([c for c in COLUNAS if c in df.columns])}/{len(COLUNAS)}</div></div>', unsafe_allow_html=True)

        st.markdown("---")
        c1, c2, c3, c4 = st.columns([1.2, 2, 1.8, 1])
        with c1:
            busca_por = st.selectbox("Buscar por", ["Cidade", "CNPJ", "Nome"], index=0)
        with c2:
            placeholder = {"Cidade":"Ex: Abdon Batista","CNPJ":"00.000.000/0001-00","Nome":"Ex: Municipio de..."}[busca_por]
            identificador = st.text_input(busca_por, placeholder=placeholder)
        with c3:
            valor_str = st.text_input("Valor alvo (R$)", placeholder="Ex: 31.452,88")
        with c4:
            st.markdown("<br>", unsafe_allow_html=True)
            executar = st.button("Conciliar →", use_container_width=True, type="primary")

        if executar:
            if not identificador:
                st.error("Informe o identificador."); st.stop()
            try:
                valor = float(valor_str.replace(".", "").replace(",", "."))
            except:
                st.error("Valor alvo inválido."); st.stop()

            bp = busca_por.upper()
            if bp == "CNPJ":
                limpo = identificador.replace(".","").replace("/","").replace("-","").lstrip("0")
                col_c = "cnpj_limpo" if "cnpj_limpo" in df.columns else "cnpj"
                df_cli = df[df[col_c].str.lstrip("0") == limpo]
                escopo = "direto"
            elif bp == "CIDADE":
                df_cli = df[df["cidade"].str.strip().str.upper() == identificador.strip().upper()]
                escopo = "cidade"
            else:
                df_cli = df[df["nome"].str.strip().str.upper().str.contains(identificador.strip().upper(), na=False)]
                escopo = "direto"

            if df_cli.empty:
                st.error(f"Nenhum registro para {busca_por}: **{identificador}**"); st.stop()

            with st.spinner(f"Buscando em {len(df_cli)} nota(s)..."):
                resultados, meta = conciliar(df_cli, valor, escopo=escopo)

            st.markdown("---")
            s1,s2,s3,s4 = st.columns(4)
            s1.markdown(f'<div class="stat-box"><div class="stat-label">Notas</div><div class="stat-value">{len(df_cli)}</div></div>', unsafe_allow_html=True)
            cor_c = "yellow" if len(resultados)>1 else "green"
            s2.markdown(f'<div class="stat-box"><div class="stat-label">Combinações</div><div class="stat-value {cor_c}">{len(resultados)}</div></div>', unsafe_allow_html=True)
            s3.markdown(f'<div class="stat-box"><div class="stat-label">Tempo</div><div class="stat-value">{meta["tempo_ms"]}ms</div></div>', unsafe_allow_html=True)
            s4.markdown(f'<div class="stat-box"><div class="stat-label">Algoritmo</div><div class="stat-value" style="font-size:14px;">{meta.get("camada","—")}</div></div>', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            if bp == "CIDADE" and resultados:
                origem = meta.get("origem", "cnpj")
                label, cor, descricao = ORIGEM_CONFIG.get(origem, ("", "gray", ""))
                if descricao:
                    css_class = {"purple":"alert-escopo","yellow":"alert-multi","red":"alert-retencao"}.get(cor, "alert-multi")
                    st.markdown(f'<div class="{css_class}"><b>{label}</b> — {descricao}<br><span style="opacity:0.7;font-size:11px;">{meta.get("escopo_descricao","")}</span></div>', unsafe_allow_html=True)

            if not resultados:
                st.error("Nenhuma combinação encontrada.")
            else:
                if len(resultados) > 1:
                    st.markdown(f'<div class="alert-multi">⚡ <b>{len(resultados)} composições distintas</b>. A #1 é a mais simples.</div>', unsafe_allow_html=True)
                for i, combo in enumerate(resultados):
                    exibir_combo(combo, i, len(resultados))

    else:
        st.info("⬆️ Carregue a planilha de títulos vencidos para iniciar.")
        st.markdown("""
**Como usar:**
1. Upload da planilha `.xlsx`
2. Selecione busca (Cidade, CNPJ ou Nome)
3. Digite o identificador e o valor recebido
4. Clique em **Conciliar →**

**Destaque por célula:** Verde só nas colunas efetivamente usadas (Saldo, IR, ISS).
Se o pagador deduziu só IR, apenas Saldo e IR ficam verdes.

**O motor identifica automaticamente:**
- ✅ Composição mais simples para o valor
- ⚠️ Retenções indevidas de IR/ISS
- ⚡ Múltiplas composições quando existirem
        """)

# ══════════════════════════════════════════════════════════════
# ABA 2 — CONCILIAÇÃO EM MASSA BB
# ══════════════════════════════════════════════════════════════
with aba_massa:
    render_aba_massa_bb()

# ══════════════════════════════════════════════════════════════
# ABA 3 — CALCULADORA
# ══════════════════════════════════════════════════════════════
with aba_calc:
    aba_calculadora()

# ══════════════════════════════════════════════════════════════
# ABA 4 — LIQUIDAÇÃO POR RETENÇÃO
# ══════════════════════════════════════════════════════════════
with aba_ret:
    aba_retencao()

# ══════════════════════════════════════════════════════════════
# ABA 5 — CONCILIAÇÃO ITAÚ
# ══════════════════════════════════════════════════════════════
with aba_itau:
    render_aba_itau()

# ══════════════════════════════════════════════════════════════
# ABA 6 — CONCILIAÇÃO BB
# ══════════════════════════════════════════════════════════════
with aba_bb:
    render_aba_bb()
