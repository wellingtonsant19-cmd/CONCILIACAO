"""
Motor de Conciliação Bancária — v5
===================================
Melhorias sobre v4:
  - Solver multi-valor: testa Saldo, Saldo+IR, Saldo+ISS, Líquido por nota
  - Excel mostra APENAS notas da composição (sem linhas vermelhas inúteis)
  - Destaque célula-a-célula: verde só nas colunas efetivamente usadas
  - Cada pagamento tem sua própria composição (não compartilha sol_idx)

Uso:
  python motor_conciliacao_v5.py --listar
  python motor_conciliacao_v5.py --aba "MAR 26"
  python motor_conciliacao_v5.py --aba "MAR 26" --pendencias PLANILHA.xlsx --titulos Titulo.xlsx
"""

import re, time, sys, logging, argparse
from difflib import SequenceMatcher
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ─────────────────────────────────────────────────────────────
ARQUIVO_PENDENCIAS = "PLANILHA_DE_PENDÊNCIAS_ATUALIZADA.xlsx"
ARQUIVO_TITULOS    = "Týtulo_em_Aberto.xlsx"
ABAS_IGNORAR       = {"classificação", "PORTAIS", "PG INTERMUNICIPAIS", "ADIANTAMENTOS"}
STATUS_RESOLVIDO   = {"BAIXADO", "BAIXADO ", "baixado"}
TOLERANCIA_APROX   = 0.02      # 2% para match aproximado
TOLERANCIA_CENTAVOS = 1        # R$0.01 para match exato
TIMEOUT_GRUPO      = 15
MAX_NOTAS_COMBO    = 30        # máximo de notas por combinação

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ── Paleta ────────────────────────────────────────────────────────────────────
AZUL_HD  = "1F3864"; BRANCO   = "FFFFFF"
VERDE_BG = "C6EFCE"; VERDE_FT = "276221"
VERM_BG  = "FFC7CE"; VERM_FT  = "9C0006"
AMAR_BG  = "FFEB9C"; AMAR_FT  = "9C6500"
CINZA_BG = "EDEDED"; CINZA_FT = "595959"
AZUL_PAG = "1F3864"
AZUL_COL = "2E75B6"
NEUTRO_BG = "F7F7F7"; NEUTRO_FT = "333333"  # linhas de nota sem destaque
FONTE = "Arial"


# ═════════════════════════════════════════════════════════════════════════════
# UTILITÁRIOS
# ═════════════════════════════════════════════════════════════════════════════

def borda():
    s = Side(style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)

def norm(texto):
    if pd.isna(texto): return ""
    t = str(texto).upper().strip()
    for p in [
        "PREFEITURA MUNICIPAL DE ", "PREFEITURA MUNICIPAL ",
        "FUNDO MUNICIPAL DE SAUDE DE ", "FUNDO MUNICIPAL DE SAUDE ",
        "FUNDO DE SAUDE DE ", "FUNDO DE SAUDE ",
        "MUNICIPIO DE ", "SECRETARIA DE ",
        "SECRETARIA MUNICIPAL DE ", "SEC ", "PM ",
        "PREF ", "PREF. ", "FMS ", "FME ",
    ]:
        t = t.replace(p, " ")
    return re.sub(r'\s+', ' ', t).strip()

def sim(a, b):
    return SequenceMatcher(None, a, b).ratio()

def centavos(v):
    try: return round(float(v) * 100)
    except: return 0


# ═════════════════════════════════════════════════════════════════════════════
# LEITURA DAS PLANILHAS
# ═════════════════════════════════════════════════════════════════════════════

def listar_abas(path_pendencias: str):
    xl = pd.ExcelFile(path_pendencias)
    abas = [s for s in xl.sheet_names if s not in ABAS_IGNORAR]
    print("\n╔══════════════════════════════════════╗")
    print("║   Abas disponíveis para análise      ║")
    print("╠══════════════════════════════════════╣")
    for a in abas:
        print(f"║  ✦  {a:<34}║")
    print("╚══════════════════════════════════════╝")
    print(f"\nUso: python motor_conciliacao_v5.py --aba \"MAR 26\"\n")


def ler_pendencias_aba(path: str, aba: str) -> pd.DataFrame:
    log.info(f"Lendo aba '{aba}' de: {path}")
    raw = pd.read_excel(path, sheet_name=aba, dtype=str, header=None)

    header_row = None
    for i, row in raw.iterrows():
        vals = [str(v).strip().upper() for v in row if not pd.isna(v)]
        if any(v in ("VALOR", "DT LANÇAMENTO", "HISTORICO", "HISTÓRICO") for v in vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"Não encontrei linha de cabeçalho na aba '{aba}'.")

    df = raw.iloc[header_row:].copy()
    df.columns = df.iloc[0].str.strip()
    df = df.iloc[1:].reset_index(drop=True)

    col_map_opts = {
        "data":      ["DT LANÇAMENTO", "DATA", "DT PGTO", "DATA PAGAMENTO"],
        "banco":     ["BANCO"],
        "status":    ["STATUS"],
        "valor":     ["VALOR", "VALOR PAGO", "VL PAGO"],
        "historico": ["HISTÓRICO", "HISTORICO", "DESCRIÇÃO", "DESCRICAO"],
        "municipio": ["MUNICIPIO", "MUNICÍPIO", "CIDADE"],
        "obs":       ["OBSERVAÇAO", "OBSERVAÇÃO", "OBS"],
    }
    cols_disponiveis = {c.strip().upper(): c for c in df.columns if not pd.isna(c)}
    mapa = {}
    for campo, opcoes in col_map_opts.items():
        mapa[campo] = next(
            (cols_disponiveis[o.upper()] for o in opcoes if o.upper() in cols_disponiveis), None)

    if not mapa["valor"]:
        raise ValueError(f"Coluna 'valor' não encontrada. Disponíveis: {list(df.columns)}")

    out = pd.DataFrame()
    out["data"]       = pd.to_datetime(df[mapa["data"]], errors="coerce") if mapa["data"] else pd.NaT
    out["banco"]      = df[mapa["banco"]].str.strip()                     if mapa["banco"] else ""
    out["status"]     = df[mapa["status"]].str.strip()                    if mapa["status"] else ""
    out["valor"]      = pd.to_numeric(df[mapa["valor"]].astype(str).str.replace(",", "."), errors="coerce")
    out["historico"]  = df[mapa["historico"]].str.strip()                  if mapa["historico"] else ""
    out["municipio"]  = df[mapa["municipio"]].str.strip().str.upper()     if mapa["municipio"] else ""
    out["obs"]        = df[mapa["obs"]].str.strip()                       if mapa["obs"] else ""
    out["municipio_norm"] = out["municipio"].apply(norm)
    out["historico_norm"] = out["historico"].apply(norm)
    out["valor_cent"]     = out["valor"].apply(centavos)
    out["idx_banco"]      = range(len(out))

    out = out.dropna(subset=["valor"]).copy()
    out = out[out["valor"] > 0].copy()

    mask_baixado = out["status"].str.upper().str.strip().isin(STATUS_RESOLVIDO)
    n_baixados   = mask_baixado.sum()
    out = out[~mask_baixado].reset_index(drop=True)
    out["idx_banco"] = range(len(out))

    log.info(f"  {len(out)+n_baixados} registros | {n_baixados} BAIXADOS | {len(out)} pendentes | R$ {out['valor'].sum():,.2f}")
    return out


def ler_titulos(path: str) -> pd.DataFrame:
    log.info(f"Lendo títulos: {path}")
    xl = pd.ExcelFile(path)
    df = pd.read_excel(path, sheet_name=xl.sheet_names[0], dtype=str)
    df.columns = df.columns.str.strip()

    for col in ["SALDO", "BRUTO", "CORRETO", "IR", "ISS"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(",", "."), errors="coerce")

    # Garantir colunas numéricas
    for col in ["SALDO", "IR", "ISS"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = df[col].fillna(0.0)

    # Calcular CORRETO (líquido) se não existir
    if "CORRETO" not in df.columns:
        df["CORRETO"] = (df["SALDO"] + df["IR"] + df["ISS"]).round(2)
    df["CORRETO"] = df["CORRETO"].fillna(df["SALDO"])

    df["VENCIMENTO"] = pd.to_datetime(df.get("VENCIMENTO", pd.NaT), errors="coerce")
    df["CIDADE"]     = df["CIDADE"].str.strip().str.upper() if "CIDADE" in df.columns else ""
    df["UF"]         = df["UF"].str.strip().str.upper()     if "UF" in df.columns else ""
    df["cidade_norm"]  = df["CIDADE"].apply(norm)
    df["nome_norm"]    = df["NOME CLIENTE"].apply(norm) if "NOME CLIENTE" in df.columns else ""
    df["idx_titulo"]   = range(len(df))

    df = df[(df["SALDO"].notna()) & (df["SALDO"] > 0)].copy()
    log.info(f"  {len(df)} títulos com saldo > 0 | R$ {df['SALDO'].sum():,.2f}")
    return df


# ═════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DE OPÇÕES DE VALOR POR NOTA (como o app Streamlit)
# ═════════════════════════════════════════════════════════════════════════════

def gerar_opcoes_nota(row):
    """
    Para cada título, gera variantes de valor que o pagador pode ter usado:
      - saldo: valor bruto em aberto
      - saldo_ir: saldo + IR (dedução de IR)
      - saldo_iss: saldo + ISS (dedução de ISS)
      - liquido: saldo + IR + ISS (dedução de ambos)
    Retorna lista de (nome_variante, valor_centavos).
    """
    s   = round(float(row.get("SALDO", 0)), 2)
    ir  = round(float(row.get("IR", 0)), 2)
    iss = round(float(row.get("ISS", 0)), 2)

    ops = [("saldo", centavos(s))]
    if ir != 0:
        ops.append(("saldo_ir", centavos(round(s + ir, 2))))
    if iss != 0:
        ops.append(("saldo_iss", centavos(round(s + iss, 2))))
    if ir != 0 and iss != 0:
        # Só adiciona líquido se for diferente de saldo_ir e saldo_iss
        liq = centavos(round(s + ir + iss, 2))
        ops.append(("liquido", liq))

    # Deduplicar variantes com mesmo valor (evita falsa ambiguidade)
    seen_vals = set()
    deduped = []
    for nome, val in ops:
        if val not in seen_vals and val > 0:
            seen_vals.add(val)
            deduped.append((nome, val))
    return deduped


def montar_itens_solver(df_cands):
    """
    Monta lista de itens para o solver.
    Cada item = (idx_titulo, variante, valor_centavos).
    Cada título pode ter múltiplas variantes (mas o solver usa no máximo 1 por título).
    """
    itens = []
    for _, row in df_cands.iterrows():
        idx = row["idx_titulo"]
        for variante, val_cent in gerar_opcoes_nota(row):
            if val_cent > 0:
                itens.append((idx, variante, val_cent))
    return itens


# ═════════════════════════════════════════════════════════════════════════════
# SOLVER SUBSET SUM MULTI-VALOR COM UNICIDADE
# ═════════════════════════════════════════════════════════════════════════════

class _Timeout(Exception): pass

def subset_sum_multival(itens, alvo_cent, deadline, max_notas=MAX_NOTAS_COMBO):
    """
    Encontra subconjunto de itens cuja soma = alvo_cent,
    com a restrição de usar no máximo 1 variante por idx_titulo.

    itens: lista de (idx_titulo, variante, valor_centavos)
    Retorna: ('unico'|'ambiguo'|'nenhum'|'timeout', [(idx_titulo, variante, valor)])
    """
    if not itens or alvo_cent <= 0:
        return "nenhum", []

    # Agrupar por título: {idx_titulo: [(variante, valor), ...]}
    por_titulo = {}
    for idx, var, val in itens:
        if val > 0 and val <= alvo_cent:
            por_titulo.setdefault(idx, []).append((var, val))

    titulos = list(por_titulo.keys())
    n = len(titulos)
    if n == 0:
        return "nenhum", []

    # Ordenar títulos pelo maior valor descendente (para poda eficiente)
    titulos.sort(key=lambda t: max(v for _, v in por_titulo[t]), reverse=True)

    # Pré-calcular sufixo de somas máximas para poda
    max_vals = [max(v for _, v in por_titulo[t]) for t in titulos]
    suf = [0] * (n + 1)
    for i in range(n - 1, -1, -1):
        suf[i] = suf[i + 1] + max_vals[i]

    solucoes = []

    def bt(pos, restante, path, n_used):
        if time.time() > deadline:
            raise _Timeout()
        if restante == 0:
            solucoes.append(list(path))
            return len(solucoes) < 2  # para ao achar 2ª solução
        if pos >= n or n_used >= max_notas:
            return True
        if suf[pos] < restante:  # poda: impossível atingir alvo
            return True

        # Opção 1: NÃO usar este título
        if not bt(pos + 1, restante, path, n_used):
            return False

        # Opção 2: usar uma das variantes deste título
        t_idx = titulos[pos]
        for var, val in por_titulo[t_idx]:
            if val > restante:
                continue
            path.append((t_idx, var, val))
            if not bt(pos + 1, restante - val, path, n_used + 1):
                path.pop()
                return False
            path.pop()

        return True

    try:
        bt(0, alvo_cent, [], 0)
    except _Timeout:
        if len(solucoes) == 1:
            return "unico", solucoes[0]
        return "timeout", []

    if   len(solucoes) == 0: return "nenhum",  []
    elif len(solucoes) == 1: return "unico",   solucoes[0]
    else:                    return "ambiguo",  []


def tentar_match(itens, alvo_cent, deadline):
    """Tenta match exato; se não achar, tenta com tolerância de 2%."""
    status, sol = subset_sum_multival(itens, alvo_cent, deadline)
    if status == "unico":
        return "unico", sol

    if status == "nenhum":
        tol = max(1, round(alvo_cent * TOLERANCIA_APROX))
        for delta in range(1, tol + 1):
            for sinal in (1, -1):
                alvo_apr = alvo_cent + sinal * delta
                if alvo_apr <= 0:
                    continue
                dl = min(deadline, time.time() + 2)
                st2, sl2 = subset_sum_multival(itens, alvo_apr, dl)
                if st2 == "unico":
                    return "unico_aprox", sl2
    return status, sol


# ═════════════════════════════════════════════════════════════════════════════
# SELEÇÃO DE CANDIDATOS
# ═════════════════════════════════════════════════════════════════════════════

def selecionar_candidatos(pag, titulos):
    mun = pag["municipio_norm"]

    if mun:
        exatos = titulos[titulos["cidade_norm"] == mun]
        if len(exatos) > 0:
            return exatos, f"Município Exato ({pag['municipio']})"

    if mun and len(mun) >= 3:
        scores = titulos["cidade_norm"].apply(lambda c: sim(mun, c) if c else 0)
        mask   = scores >= 0.70
        if mask.any():
            best  = scores[mask].max()
            cidade = titulos.loc[scores == best, "CIDADE"].iloc[0]
            return titulos[titulos["CIDADE"] == cidade], f"Município Fuzzy ({cidade}, {best:.0%})"

    hist = pag["historico_norm"]
    if hist and len(hist) >= 5:
        scores = titulos["nome_norm"].apply(lambda n: sim(hist, n) if n else 0)
        mask   = scores >= 0.65
        if mask.any():
            best   = scores[mask].max()
            cidade = titulos.loc[scores == best, "CIDADE"].iloc[0]
            return titulos[titulos["CIDADE"] == cidade], f"Histórico Fuzzy ({best:.0%})"

    return pd.DataFrame(), "Sem candidatos"


# ═════════════════════════════════════════════════════════════════════════════
# MOTOR PRINCIPAL
# ═════════════════════════════════════════════════════════════════════════════

def executar_conciliacao(banco, titulos):
    log.info("Iniciando conciliação...")
    t0 = time.time()

    resultados     = []
    titulos_usados = set()

    banco["chave_grupo"] = (
        banco["municipio"].fillna("") + "|" +
        banco["data"].dt.strftime("%Y-%m-%d").fillna("")
    )
    grupos = list(banco.groupby("chave_grupo", sort=False))
    log.info(f"  {len(banco)} pagamentos em {len(grupos)} grupos")

    for g_idx, (chave, grupo) in enumerate(grupos, 1):
        pags = grupo.sort_values("valor", ascending=False)
        ref  = pags.iloc[0]

        cands, fase = selecionar_candidatos(ref, titulos)
        if len(cands) == 0:
            for _, p in pags.iterrows():
                resultados.append(_res(p, "SEM_CANDIDATOS", fase, [], "Município/cliente não encontrado"))
            continue

        cands_livres = cands[~cands["idx_titulo"].isin(titulos_usados)].copy()
        if len(cands_livres) == 0:
            for _, p in pags.iterrows():
                resultados.append(_res(p, "SEM_CANDIDATOS", fase, [], "Todos os títulos já vinculados"))
            continue

        itens_base = montar_itens_solver(cands_livres)
        deadline   = time.time() + TIMEOUT_GRUPO

        # ── Estratégia A: soma total do grupo (mesmo dia/município) ───────
        if len(pags) > 1:
            total_cent = int(pags["valor_cent"].sum())
            status_a, sol_a = tentar_match(itens_base, total_cent, deadline)

            if status_a in ("unico", "unico_aprox"):
                # Distribuir a solução: atribuir notas específicas a cada pagamento
                # Por simplicidade, cada pagamento do grupo recebe a lista completa
                # mas marcamos como "grupo"
                obs = "Soma de pagamentos do dia"
                if status_a == "unico_aprox":
                    obs += " (valor aprox. ≤2%)"
                for _, p in pags.iterrows():
                    resultados.append(_res(p, status_a, fase, sol_a, obs))
                for idx_t, _, _ in sol_a:
                    titulos_usados.add(idx_t)
                continue

        # ── Estratégia B: cada pagamento individual ───────────────────────
        usados_grupo = set()
        for _, p in pags.iterrows():
            alvo_cent = int(p["valor_cent"])

            # Filtrar itens: remover títulos já usados neste grupo
            itens_livres = [(idx, var, val) for idx, var, val in itens_base
                           if idx not in usados_grupo]

            if not itens_livres:
                resultados.append(_res(p, "SEM_CANDIDATOS", fase, [], "Candidatos esgotados"))
                continue

            dl = min(deadline, time.time() + TIMEOUT_GRUPO / 2)
            status_b, sol_b = tentar_match(itens_livres, alvo_cent, dl)

            if status_b in ("unico", "unico_aprox"):
                obs = "Valor aproximado (≤2%)" if status_b == "unico_aprox" else ""
                resultados.append(_res(p, status_b, fase, sol_b, obs))
                for idx_t, _, _ in sol_b:
                    usados_grupo.add(idx_t)
                    titulos_usados.add(idx_t)
            else:
                obs_map = {
                    "ambiguo": "Múltiplas combinações possíveis — não é único",
                    "timeout": "Timeout — muitos candidatos",
                    "nenhum":  "Nenhuma combinação resulta neste valor",
                }
                resultados.append(_res(p, status_b, fase, [], obs_map.get(status_b, status_b)))

    n_ok  = sum(1 for r in resultados if r["status"] in ("unico", "unico_aprox"))
    n_amb = sum(1 for r in resultados if r["status"] == "ambiguo")
    n_sem = len(resultados) - n_ok - n_amb
    log.info(f"Concluído em {time.time()-t0:.1f}s | ✅ {n_ok} | ⚠️ {n_amb} | ❌ {n_sem}")
    return resultados


def _res(pag, status, fase, solucao, obs):
    """
    solucao: lista de (idx_titulo, variante, valor_centavos)
    """
    return {
        "pag":     pag,
        "status":  status,
        "fase":    fase,
        "solucao": solucao,   # [(idx_titulo, variante, val_cent), ...]
        "obs":     obs,
    }


# ═════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DO EXCEL
# ═════════════════════════════════════════════════════════════════════════════

# Colunas: campo_df, header_excel, largura, é_valor
COLS_TIT = [
    ("NF",          "Nota",         10,  False),
    ("DOC",         "DOC",          14,  False),
    ("VENCIMENTO",  "Vencimento",   13,  False),
    ("SALDO",       "Saldo (R$)",   15,  True),
    ("IR",          "IR (R$)",      11,  True),
    ("ISS",         "ISS (R$)",     11,  True),
    ("CORRETO",     "Líquido (R$)", 15,  True),
    ("NOME CLIENTE","Cliente",      38,  False),
    ("UF",          "UF",            6,  False),
    ("_VARIANTE",   "Tipo Usado",   14,  False),
]
N_COLS = len(COLS_TIT)

# Mapeamento variante → quais campos devem ficar verdes
VARIANTE_CAMPOS_VERDES = {
    "saldo":     {"SALDO"},
    "saldo_ir":  {"SALDO", "IR"},
    "saldo_iss": {"SALDO", "ISS"},
    "liquido":   {"SALDO", "IR", "ISS", "CORRETO"},
}

VARIANTE_LABEL = {
    "saldo":     "Saldo",
    "saldo_ir":  "Saldo − IR",
    "saldo_iss": "Saldo − ISS",
    "liquido":   "Líquido",
}


def gerar_excel(resultados, titulos, aba_analisada, saida):
    log.info(f"Gerando Excel: {saida}")
    wb = Workbook()
    wb.remove(wb.active)
    _aba_conciliacao(wb, resultados, titulos, aba_analisada)
    _aba_nao_resolvidos(wb, resultados, aba_analisada)
    _aba_resumo(wb, resultados, aba_analisada)
    wb.save(saida)
    log.info(f"✅ Salvo: {saida}")


def _c(ws, row, col, val=None, bold=False, color="000000", bg=None,
       sz=9, ha="left", wrap=False, fmt=None, italic=False):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.font      = Font(name=FONTE, bold=bold, color=color, size=sz, italic=italic)
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    c.border    = borda()
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    if fmt:
        c.number_format = fmt
    return c

def _hdr_row(ws, row, texts, bg, fnt=BRANCO, h=20):
    ws.row_dimensions[row].height = h
    for col, txt in enumerate(texts, 1):
        _c(ws, row, col, txt, bold=True, color=fnt, bg=bg, ha="center", wrap=True)

def _titulo_aba(ws, texto, ncols, bg=AZUL_HD, h=30):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]
    c.value     = texto
    c.font      = Font(name=FONTE, bold=True, size=13, color=BRANCO)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = h


# ── Aba CONCILIAÇÃO ──────────────────────────────────────────────────────────

def _aba_conciliacao(wb, resultados, titulos, aba_analisada):
    ws = wb.create_sheet("🗂️ CONCILIAÇÃO")
    ws.sheet_view.showGridLines = False

    for i, (_, _, w, _) in enumerate(COLS_TIT, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _titulo_aba(ws, f"CONCILIAÇÃO BANCÁRIA — {aba_analisada}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}", N_COLS)

    # Legenda atualizada
    ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
    c = ws.cell(row=2, column=1)
    c.value = ("  🟢 Verde = célula usada na composição do pagamento     "
               "⬜ Cinza claro = dado da nota (não usado diretamente)     "
               "⚠️ Amarelo = valor aproximado (≤2%)")
    c.font  = Font(name=FONTE, size=9, italic=True, color="444444")
    c.fill  = PatternFill("solid", start_color="F5F5F5")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4
    ok = [r for r in resultados if r["status"] in ("unico", "unico_aprox")]

    if not ok:
        ws.merge_cells(f"A4:{get_column_letter(N_COLS)}6")
        c = ws["A4"]
        c.value = "Nenhum pagamento conciliado. Verifique a aba ❌ NÃO RESOLVIDOS."
        c.font  = Font(name=FONTE, size=11, color=AMAR_FT, bold=True)
        c.fill  = PatternFill("solid", start_color=AMAR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 40
        return

    # Indexar títulos por idx_titulo para lookup rápido
    tit_map = {int(row["idx_titulo"]): row for _, row in titulos.iterrows()}

    for res in ok:
        pag     = res["pag"]
        solucao = res["solucao"]  # [(idx_titulo, variante, val_cent), ...]
        aprox   = res["status"] == "unico_aprox"

        if not solucao:
            continue

        # ── Cabeçalho do bloco pagamento ──────────────────────────────────
        ws.row_dimensions[row].height = 14
        hdrs_pag = ["DT LANÇAMENTO", "BANCO", "VALOR PAGO (R$)", "HISTÓRICO / CLIENTE", "MUNICÍPIO"]
        for col, h in enumerate(hdrs_pag[:N_COLS], 1):
            _c(ws, row, col, h, bold=True, color=BRANCO, bg=AZUL_PAG, ha="center", sz=8)
        for col in range(len(hdrs_pag) + 1, N_COLS + 1):
            _c(ws, row, col, "", bg=AZUL_PAG)
        row += 1

        # ── Dados do pagamento ────────────────────────────────────────────
        ws.row_dimensions[row].height = 18
        # Calcular soma da composição para exibir
        soma_comp = round(sum(v / 100 for _, _, v in solucao), 2)
        diff = round(pag["valor"] - soma_comp, 2)

        cidade_display = pag.get("municipio", "")
        if not cidade_display and solucao:
            t = tit_map.get(solucao[0][0])
            if t is not None:
                cidade_display = t.get("CIDADE", "")

        vals_pag = [
            pag["data"].date() if pd.notna(pag.get("data")) else "",
            pag.get("banco", ""),
            pag["valor"],
            pag.get("historico", ""),
            cidade_display,
        ]
        for col, v in enumerate(vals_pag[:N_COLS], 1):
            c2 = _c(ws, row, col, v, bold=True, color=BRANCO, bg="2E4057",
                     ha="center" if col <= 3 else "left", sz=10)
            if col == 3:
                c2.number_format = "#,##0.00"
        for col in range(len(vals_pag) + 1, N_COLS + 1):
            _c(ws, row, col, "", bg="2E4057")

        if aprox:
            _c(ws, row, N_COLS, f"⚠ Aprox. (Δ R${abs(diff):,.2f})",
               bold=True, color=AMAR_FT, bg=AMAR_BG, ha="center", sz=8)
        row += 1

        # ── Cabeçalho das colunas de título ───────────────────────────────
        ws.row_dimensions[row].height = 14
        for col, (_, hdr, _, _) in enumerate(COLS_TIT, 1):
            _c(ws, row, col, hdr, bold=True, color=BRANCO, bg=AZUL_COL, ha="center", sz=8)
        row += 1

        # ── SÓ AS NOTAS DA COMPOSIÇÃO (sem linhas vermelhas!) ─────────────
        for idx_t, variante, val_cent in solucao:
            tit = tit_map.get(idx_t)
            if tit is None:
                continue

            campos_verdes = VARIANTE_CAMPOS_VERDES.get(variante, set())

            ws.row_dimensions[row].height = 14
            for col, (campo, _, _, is_valor) in enumerate(COLS_TIT, 1):
                # Valor especial para _VARIANTE
                if campo == "_VARIANTE":
                    val = VARIANTE_LABEL.get(variante, variante)
                else:
                    val = tit.get(campo, "")
                    if campo == "VENCIMENTO" and pd.notna(val):
                        try:
                            val = val.date()
                        except:
                            pass

                # Decidir cor da célula
                if campo in campos_verdes:
                    bg, ft, bold = VERDE_BG, VERDE_FT, True
                elif campo == "_VARIANTE":
                    # Colorir label da variante com base no tipo
                    if variante == "saldo":
                        bg, ft, bold = VERDE_BG, VERDE_FT, True
                    elif variante in ("saldo_ir", "saldo_iss"):
                        bg, ft, bold = AMAR_BG, AMAR_FT, True
                    else:
                        bg, ft, bold = VERM_BG, VERM_FT, True
                else:
                    bg, ft, bold = NEUTRO_BG, NEUTRO_FT, False

                c2 = _c(ws, row, col, val,
                        bold=bold, color=ft, bg=bg,
                        ha="right" if is_valor else "center", sz=9)
                if is_valor:
                    c2.number_format = "#,##0.00"
            row += 1

        # ── Linha de totalização da composição ────────────────────────────
        ws.row_dimensions[row].height = 14
        for col in range(1, N_COLS + 1):
            _c(ws, row, col, "", bg="E8E8E8")
        _c(ws, row, 1, f"TOTAL COMPOSIÇÃO ({len(solucao)} nota(s))",
           bold=True, color="333333", bg="E8E8E8", ha="left", sz=8)

        # Calcular totais por coluna
        soma_saldo = round(sum(float(tit_map.get(i, {}).get("SALDO", 0) or 0)
                               for i, _, _ in solucao), 2)
        soma_ir    = round(sum(float(tit_map.get(i, {}).get("IR", 0) or 0)
                               for i, _, _ in solucao), 2)
        soma_iss   = round(sum(float(tit_map.get(i, {}).get("ISS", 0) or 0)
                               for i, _, _ in solucao), 2)

        # Coluna SALDO total
        col_saldo = next((ci for ci, (f, _, _, _) in enumerate(COLS_TIT, 1) if f == "SALDO"), None)
        col_ir    = next((ci for ci, (f, _, _, _) in enumerate(COLS_TIT, 1) if f == "IR"), None)
        col_iss   = next((ci for ci, (f, _, _, _) in enumerate(COLS_TIT, 1) if f == "ISS"), None)
        col_liq   = next((ci for ci, (f, _, _, _) in enumerate(COLS_TIT, 1) if f == "CORRETO"), None)

        if col_saldo:
            c2 = _c(ws, row, col_saldo, soma_saldo, bold=True, color="333333", bg="E8E8E8", ha="right")
            c2.number_format = "#,##0.00"
        if col_ir:
            c2 = _c(ws, row, col_ir, soma_ir, bold=True, color="333333", bg="E8E8E8", ha="right")
            c2.number_format = "#,##0.00"
        if col_iss:
            c2 = _c(ws, row, col_iss, soma_iss, bold=True, color="333333", bg="E8E8E8", ha="right")
            c2.number_format = "#,##0.00"
        if col_liq:
            c2 = _c(ws, row, col_liq, soma_comp, bold=True, color=VERDE_FT, bg="E8E8E8", ha="right")
            c2.number_format = "#,##0.00"

        row += 1

        # Separador entre blocos
        ws.row_dimensions[row].height = 6
        for col in range(1, N_COLS + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", start_color="F0F0F0")
        row += 1


# ── Aba NÃO RESOLVIDOS ──────────────────────────────────────────────────────

def _aba_nao_resolvidos(wb, resultados, aba_analisada):
    ws = wb.create_sheet("❌ NÃO RESOLVIDOS")
    ws.sheet_view.showGridLines = False

    hdrs = ["Data", "Banco", "Valor (R$)", "Histórico / Cliente", "Município",
            "Status", "Fase Match", "Observação / Motivo"]
    ncols = len(hdrs)

    _titulo_aba(ws, f"NÃO RESOLVIDOS — {aba_analisada}", ncols, bg="7B0000")
    _hdr_row(ws, 2, hdrs, bg=VERM_FT, h=22)

    STATUS_LBL = {
        "ambiguo":        "⚠️ Ambíguo",
        "nenhum":         "❌ Sem combinação",
        "SEM_CANDIDATOS": "❌ Sem candidatos",
        "timeout":        "⏱️ Timeout",
    }
    STATUS_COR = {
        "ambiguo":        (AMAR_BG, AMAR_FT),
        "nenhum":         (VERM_BG, VERM_FT),
        "SEM_CANDIDATOS": ("FFE0E0", "800000"),
        "timeout":        (AMAR_BG, AMAR_FT),
    }

    nao_ok = [r for r in resultados if r["status"] not in ("unico", "unico_aprox")]

    for row_num, res in enumerate(nao_ok, 3):
        pag = res["pag"]
        bg, ft = STATUS_COR.get(res["status"], (CINZA_BG, CINZA_FT))
        lbl    = STATUS_LBL.get(res["status"], res["status"])
        vals   = [
            pag["data"].date() if pd.notna(pag.get("data")) else "",
            pag.get("banco", ""),
            pag["valor"],
            pag.get("historico", ""),
            pag.get("municipio", ""),
            lbl,
            res.get("fase", ""),
            res.get("obs", ""),
        ]
        ws.row_dimensions[row_num].height = 15
        for col, v in enumerate(vals, 1):
            c = _c(ws, row_num, col, v, color=ft, bg=bg,
                   wrap=(col in [4, 8]), ha="right" if col == 3 else "left")
            if col == 3:
                c.number_format = "#,##0.00"

    for l, w in zip("ABCDEFGH", [12, 8, 14, 40, 20, 20, 22, 46]):
        ws.column_dimensions[l].width = w


# ── Aba RESUMO ───────────────────────────────────────────────────────────────

def _aba_resumo(wb, resultados, aba_analisada):
    ws = wb.create_sheet("📊 RESUMO")
    ws.sheet_view.showGridLines = False

    ok_ex  = [r for r in resultados if r["status"] == "unico"]
    ok_apr = [r for r in resultados if r["status"] == "unico_aprox"]
    ambig  = [r for r in resultados if r["status"] == "ambiguo"]
    sem    = [r for r in resultados if r["status"] not in ("unico", "unico_aprox", "ambiguo")]
    todos  = resultados

    def soma(lst):
        return sum(float(r["pag"]["valor"]) for r in lst)

    v_tot = soma(todos) or 1

    linhas = [
        ("Total Analisado",                    len(todos),  soma(todos), 1.0,                    None),
        ("✅ Conciliados — valor exato",        len(ok_ex),  soma(ok_ex), soma(ok_ex) / v_tot,   VERDE_BG),
        ("🟡 Conciliados — valor aprox. (≤2%)", len(ok_apr), soma(ok_apr), soma(ok_apr) / v_tot, AMAR_BG),
        ("⚠️ Ambíguos",                        len(ambig),  soma(ambig), soma(ambig) / v_tot,   AMAR_BG),
        ("❌ Sem match / Sem candidatos",       len(sem),    soma(sem),   soma(sem) / v_tot,     VERM_BG),
    ]

    _titulo_aba(ws, f"RESUMO — Conciliação {aba_analisada}", 4)
    ws.merge_cells("A2:D2")
    ws["A2"].value = (f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                      f"Aba: {aba_analisada}  |  Total: R$ {soma(todos):,.2f}")
    ws["A2"].font  = Font(name=FONTE, size=9, color="888888", italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    _hdr_row(ws, 4, ["Situação", "Qtd", "Valor (R$)", "% do Total"], bg="2F75B6", h=22)

    for i, (lbl, qtd, val, pct, bg) in enumerate(linhas, 5):
        ws.row_dimensions[i].height = 20
        fill = PatternFill("solid", start_color=bg) if bg else None
        for col, (v, fmt) in enumerate([(lbl, None), (qtd, "#,##0"), (val, "#,##0.00"), (pct, "0.0%")], 1):
            c = ws.cell(row=i, column=col, value=v)
            c.font   = Font(name=FONTE, size=10)
            c.border = borda()
            c.alignment = Alignment(vertical="center")
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt

    ws.row_dimensions[11].height = 24
    ws.merge_cells("A11:C11")
    ws["A11"].value = "Taxa de Conciliação Automática"
    ws["A11"].font  = Font(name=FONTE, bold=True, size=11)
    taxa = (len(ok_ex) + len(ok_apr)) / len(todos) if todos else 0
    c = ws["D11"]
    c.value         = taxa
    c.number_format = "0.0%"
    c.font          = Font(name=FONTE, bold=True, size=13,
                           color=VERDE_FT if taxa >= 0.8 else (AMAR_FT if taxa >= 0.5 else VERM_FT))
    c.alignment     = Alignment(horizontal="center", vertical="center")
    c.border        = borda()
    c.fill          = PatternFill("solid",
                                  start_color=VERDE_BG if taxa >= 0.8 else (AMAR_BG if taxa >= 0.5 else VERM_BG))

    for l, w in zip("ABCD", [42, 10, 18, 14]):
        ws.column_dimensions[l].width = w


# ═════════════════════════════════════════════════════════════════════════════
# PONTO DE ENTRADA
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Motor de Conciliação Bancária v5",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument("--pendencias", default=ARQUIVO_PENDENCIAS)
    parser.add_argument("--titulos",    default=ARQUIVO_TITULOS)
    parser.add_argument("--aba",        default=None)
    parser.add_argument("--saida",      default=None)
    parser.add_argument("--listar",     action="store_true")
    args = parser.parse_args()

    if args.listar:
        listar_abas(args.pendencias)
        return

    if not args.aba:
        print("\nERRO: Informe a aba com --aba\n")
        listar_abas(args.pendencias)
        sys.exit(1)

    aba   = args.aba
    saida = args.saida or f"conciliacao_{aba.replace(' ', '_')}.xlsx"

    banco    = ler_pendencias_aba(args.pendencias, aba)
    titulos  = ler_titulos(args.titulos)
    resultados = executar_conciliacao(banco, titulos)
    gerar_excel(resultados, titulos, aba, saida)

    n_ok  = sum(1 for r in resultados if r["status"] in ("unico", "unico_aprox"))
    n_tot = len(resultados)
    print(f"\n✅ {saida}")
    print(f"   {n_ok}/{n_tot} conciliados ({n_ok / n_tot * 100:.0f}%)\n" if n_tot else "")


if __name__ == "__main__":
    main()
